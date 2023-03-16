import math

import pandas as pd
from pandas import ExcelWriter
import os
import random
import openpyxl
import init
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet import dimensions
from openpyxl.styles import PatternFill, Alignment
from openpyxl import load_workbook
import socket
from conflict import resolveConflictSingles, resolveConflictCouples
from init import updateDanceDfs, getNode, buildInstTree
from Structures import Heat, HeatList, ConflictLog, ConflictItemSingle, ConflictItemCouple
from uuid import getnode as gma

'''
FIle with all methods used to create the Freestyle itenerary
'''

IP = ""
port = 8989

max_conflicts = 2000
def run():
    # Create socket
    client = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
    # TODO encrypt the file with the ip addr, add read and decryption here
    client.connect((IP, port))

    # Import all input Sheets independently
    file = os.getcwd().replace('\src', "") + '\FreestyleEventPlannerInput.xlsm'

    # get file and size
    bytestr = open(file, "rb")
    # fileSize = os.path.getsize(file)

    # Make file name the Event name in the server
    df_set = pd.read_excel(file, sheet_name='Settings', index_col=0)
    eventName = df_set['Data']['Event Name']

    # Get mac address, if not faked send all the data
    mac1 = gma()
    if mac1 == gma():
        # Send all info
        client.send((eventName + ".xlsm").encode())
        client.send(gma())
        # client.send(str(fileSize).encode())

        data = bytestr.read()
        client.sendall(data)
        client.send(b"<Terminate>")

    # TODO continually run until a message comes from the server, depending on the message execute differently
'''



'''
def partitionData():
    # Define base column lengths
    baseSingleCols = 8
    baseCoupleCols = 10
    baseInstructorCols = 4
    baseSettingsRows = 8
    """
    # DF's to make debugging faster
    data_set = {'Event Days'
                'Freestyle Time for Day 1 (in Hrs)':
                'Freestyle Time for Day 2 (in Hrs)':
                'Freestyle Time for Day 3 (in Hrs)':
                'Number of Judges':
                'Couples per Judge':
                'Number of Ballrooms':
                'Judges per Ballroom':
                'Heat Duration (Min)':
                'Heat Intermission (Min)':
                'Number of Age Brackets':
                'Age Bracket 1':
                'Age Bracket 2':
                'Age Bracket 3':
                'Event Name':
                'Data': 3,
                        4,
                        8,
                        10,
                        8,
                        4,
                        2,
                        4,
                        1.5,
                        1,
                        3,
                        35,
                        55,
                        56,
                        Hello World
                }
    
    df_set = pd.DataFrame(data=data_set)
    
    data_cat = {'Dance': ['Foxtrot',
                          'ChaCha',
                          'Bachata',
                          'Showdown'],
                'Genre': ['Smooth',
                          'Rhythm',
                          'Specialty',
                          'Event'],
                'Price': [35,
                          35,
                          35,
                          75]
                }
    df_cat = pd.DataFrame.from_dict(data_cat)
    data_inst = {'Dancer #': [205,
                              206],
                'First Name': ['Ashley',
                               'Jennifer'],
                'Last Name': ['Kipps',
                              'Howard'],
                'School': ['Richmond',
                           'Richmond']
                 }
    init.df_inst = pd.DataFrame.from_dict(data_inst)
    data_sing = {'Dancer #': [101,
                              102,
                              103,
                              104],
                 'First Name': ['Dalton',
                                "Marin",
                                "Kevin",
                                "Kayla"],
                 'Last Name': ['Dabney',
                               "Walters",
                               "Donahue",
                               "Dabney"],
                 'Age': [25,
                         24,
                         32,
                         23],
                 'Lead/Follow': ['Lead',
                                 "Follow",
                                 'Lead',
                                 "Follow"],
                 'Level': ["B1", "B1", "B3", "B4"],
                 "Instructor Dancer #'s": [[205,206],
                                        [206,205],
                                           [205, 206],
                                           [206]],
                 'School': ['Richmond',
                            'Richmond',
                            'Richmond',
                            'Richmond'],
                 'Open Foxtrot': [1,
                                  1,
                                  2,
                                  2],
                 'Closed ChaCha': [2,
                                   4,
                                   0,
                                   0],
                 'Showdown': [1,
                              3,
                              0,
                              0]
                 }
    df_sing = pd.DataFrame.from_dict(data_sing)
    data_coup = {'Lead Dancer #': [425],
                 'Lead First Name': ['Bill'],
                 'Lead Last Name': ['Richardson'],
                 'Lead Age': [54],
                 'Follow Dancer #': [426],
                 'Follow First Name': ['Suzi'],
                 'Follow Last Name': ['Richardson'],
                 'Follow Age': [52],
                 'Level': ["S4"],
                 'School': ['Richmond'],
                 'Open Foxtrot': [3],
                 'Closed ChaCha': [1],
                 'Closed Bachata': [2],
                 'Showdown': [1]
                 }
    df_coup = pd.DataFrame.from_dict(data_coup)
    """
    # Import all input Sheets independently
    file = os.getcwd().replace('\src', "") + '\FreestyleEventPlannerInput.xlsm'
    df_set = pd.read_excel(file, sheet_name='Settings', index_col=0, engine="openpyxl")
    df_cat = pd.read_excel(file, sheet_name='Event Categories')

    # Make list of Genres
    genrelist = list(df_cat['Genre'].unique())
    # Make list of Dances/Events
    dancelist = list(df_cat['Dance'].unique())

    heats = makeHeatDict(genrelist, df_cat)

    # Stops pandas from reading useless blank columns
    cols = []
    for i in range(len(dancelist)+baseSingleCols):
        cols.append(i)

    init.df_sing = pd.read_excel(file, sheet_name='Singles', usecols=cols)
    init.df_sing['type id'] = ''
    init.df_sing["Dancer #"] = init.df_sing['Dancer #'].astype(int)
    init.df_sing["Age"] = init.df_sing['Age'].astype(int)
    # for row, entry in df_sing.iterrows():
    #     # Retrofit the instructor list data
    #     if type(entry["Instructor Dancer #'s"]) == int:
    #         tmp = [entry["Instructor Dancer #'s"]]
    #     else:
    #         tmp = [int(x) for x in entry["Instructor Dancer #'s"].split(";")]
    #     entry["Instructor Dancer #'s"] = tmp
    #     del tmp

    # Stops pandas from reading useless blank columns
    cols = []
    for i in range(len(dancelist) + baseCoupleCols):
        cols.append(i)
    # df = pd.DataFrame
    # hello = {"Hello": df}
    # deleteEmpty(hello)
    # print(len(hello))
    init.df_coup = pd.read_excel(file, sheet_name='Couples', usecols=cols)
    init.df_coup["Lead Dancer #"] = init.df_coup['Lead Dancer #'].astype(int)
    init.df_coup["Follow Dancer #"] = init.df_coup['Follow Dancer #'].astype(int)
    init.df_coup["Lead Age"] = init.df_coup['Lead Age'].astype(int)
    init.df_coup["Follow Age"] = init.df_coup['Follow Age'].astype(int)
    init.df_coup["type id"] = "C"

    init.df_inst = pd.read_excel(file, sheet_name='Instructors')
    init.df_inst["Dancer #"] = init.df_inst['Dancer #'].astype(int)

    # Get all settings data partitioned and saved
    days = int(df_set['Data']['Event Days'])
    hours_p_day = []
    # create list holding hours of heat time per day
    for each in range(days):
        hours_p_day.append(int(df_set['Data'][each + 1]))

    # make df_Dnum
    cols = ['Lead Dancer #', "Lead First Name", "Lead Last Name"]
    coupl = init.df_coup[cols]
    coupl = coupl.rename(columns={'Lead First Name': 'First Name', 'Lead Last Name': 'Last Name',
                                  'Lead Dancer #': 'Dancer #'})
    cols = ['Follow Dancer #', "Follow First Name", "Follow Last Name"]
    coupf = init.df_coup[cols]
    coupf = coupf.rename(columns={'Follow First Name': 'First Name', 'Follow Last Name': 'Last Name',
                            'Follow Dancer #': 'Dancer #'})
    cols = ['Dancer #', "First Name", "Last Name"]
    sing = init.df_sing[cols]

    cols = ['Dancer #', "First Name", "Last Name"]
    inst = init.df_inst[cols]

    init.df_Dnum = pd.concat([coupl, coupf, sing, inst])

    judges = int(df_set['Data']['Number of Judges'])
    judge_ratio = int(df_set['Data']['Couples per Judge'])
    ballrooms = int(df_set['Data']["Number of Ballrooms"])
    ballroom_judges = int(df_set['Data']['Judges per Ballroom'])
    heat_duration = df_set['Data']['Heat Duration (Min)']
    heat_intermission = df_set['Data']['Heat Intermission (Min)']
    bracket_count = int(df_set['Data']['Age Brackets'])
    eventName = df_set['Data']['Event Name']

    age_brackets = []
    for each in range(bracket_count-1):
        age_brackets.append(int(df_set['Data'][each + days + 2]))
    age_brackets.append(1000)  # Append to make sure to get the last age bracket

    # Max number of couples on the floor for a dance assume even # of judges per ballroom
    max_dance_couples = judges * judge_ratio
    # couples_per_ballroom = int(max_dance_couples / ballrooms)  # Number of Couples on a singular ballroom
    # couples_per_ballroom = 7  # Number of Couples on a singular ballroom

    heats_p_day = []  # Holds # of heats per day
    # Calculations based on Settings data
    for count, each in enumerate(hours_p_day):
        heat_time = heat_duration + heat_intermission
        heats_p_day.append((60 / heat_time) * each)

    max_heats = sum(heats_p_day)  # Max number of heats for entire event due to time settings
    heats_p_hour = max_heats / sum(hours_p_day)

    # Find # of heats per event i.e. per entry in dancelist
    ''' 
    loop through the dance list count the entrance numbers for that event
    need to account for multi entries, when I find the sum compare that to the # of contestansts and you know how many multi entries you have.
    Use that number to find

    or 

    going line by line and making a nominal heat to be sure there are no duplicate entries

    or

    don't worry about putting a heat inside a day until after they are all made.

    For each in dancelist:
        total_heat_p_dance = (df_sing.loc[each].sum() + df_coup.loc[each].sum()) / max_dance_couples
    '''
    '''
    df_list = [df_coup, df_sing]
    for each in df_list:
        each.Level = each.Level.astype(str)
    '''
    # Slice Contestant dfs based on levels
    contestant_data = sliceDfs(init.df_sing, init.df_coup)

    '''
    Loop through all dances in the heats dictionary depth first create all heats and store them 
    inside a list at the the lowest order dictionary 
    ex: {'Rhythm': {'Closed': {"Closed ChaCha" : [Heat1, Heat2, ..., Heatn]
                               "Closed East Coast Swing" : [Heat1, Heat2, ..., Heatn]}
                    .
                    .
                    .
                  }
        }
    '''
    # Used for easy access to dfs sliced by level
    AB = 0
    FB = 1
    AS = 2
    FS = 3
    AG = 4
    FG = 5

    tot_current_heats = 0  # Current total heats made
    # For each genre
    genres = list(heats.keys())
    for each in genres:
        # For each syllabus in genre
        syllabus = list(heats[each].keys())
        for every in syllabus:
            # For each event in that syllabus event
            events = list(heats[each][every].keys())
            if events == []:
                continue
            for ev in events:
                init.ev = ev
                init.dance_dfs = {"C": {}, "S": {"Lead": {}, "Follow": {}}}
                l_keys = [AB, FB, AS, FS, AG, FG]
                dance_heat_count = 0  # current Heat count for this dance event
                tot_holes = []  # number of open spaces in the heats made to reach the maximum number per heat
# ------------------------------------------ Build dfs for Selection ---------------------------------------------------
                # Slice dfs based on participation in dance event 'ev', level, type, age
                # Add identifier column Couple dfs and reformat to be ready for heat sheet print, Singles done later
                eventrow = df_cat.loc[df_cat['Dance'] == ev].reset_index(drop=True)
                couples_per_floor = int(eventrow.loc[:, "Couples Per Floor"][0])
                acceptablecouples = int(couples_per_floor/2)+2
                div = eventrow.loc[:, 'Event Divisions'][0]
                if type(div) == float:  # if blank
                    div = ["n"]
                else:
                    div = eventrow.loc[:, 'Event Divisions'][0].split(";")

                # Use Combine Age Brackets field to set up dance dfs tree
                eventages_s = []
                combineage_s = eventrow.loc[:, 'Combine Age Brackets'][0]
                if type(combineage_s) == float:  # if blank
                    eventages_s = age_brackets.copy()
                else:
                    combineage = eventrow.loc[:, 'Combine Age Brackets'][0].split(";")
                    eventage = []
                    for age in age_brackets:
                        eventage.append([age])
                    for combo in combineage:
                        first = int(combo[0])
                        last = int(combo[2])
                        if first > last:  # Swap them if entered incorrectly
                            first = last
                            last = first
                        if first > bracket_count or last > bracket_count:
                            raise Exception("Combine Age Brackets for", ev, "has a number outside", "1-" + str(bracket_count))
                        for j in range(last-first):
                            eventage[first+j-1].clear()
                    for entry in eventage:
                        if entry != []:
                            eventages_s.append(entry[0])
                # Use Combine Age Brackets field to set up dance dfs tree
                eventages_c = []
                combineage_c = eventrow.loc[:, 'Combine Age Brackets'][0]
                if type(combineage_c) == float:  # if blank
                    eventages_c = age_brackets.copy()
                else:
                    combineage = eventrow.loc[:, 'Combine Age Brackets'][0].split(";")
                    eventage = []
                    for age in age_brackets:
                        eventage.append([age])
                    for combo in combineage:
                        first = int(combo[0])
                        last = int(combo[2])
                        if first > last:  # Swap them if entered incorrectly
                            first = last
                            last = first
                        if first > bracket_count or last > bracket_count:
                            raise Exception("Combine Age Brackets for", ev, "has a number outside", "1-" + str(bracket_count))
                        for j in range(last - first):
                            eventage[first + j - 1].clear()
                    for entry in eventage:
                        if entry != []:
                            eventages_c.append(entry[0])

                # Use Combine Levels field to set up dance dfs tree for Couples
                eventlvls_c = []
                eventlvlnames_c = []
                combinelvls_c = eventrow.loc[:, 'Combine Levels'][0]
                if type(combinelvls_c) == float:  # if blank
                    eventlvlnames_c = init.lvls.copy()
                else:
                    combinelvl = eventrow.loc[:, 'Combine Levels'][0].split(";")
                    lvlnames_c = []
                    lvlcombos_c = []
                    firsts = []
                    lasts = []
                    for lvl in init.lvls:
                        lvlnames_c.append([lvl])
                    for dfs in contestant_data["Couple"]:
                        lvlcombos_c.append([dfs])
                    for combo in combinelvl:
                        concat_data_c = []
                        first = combo[0:2]
                        firsts.append(first)
                        first = init.lvl_conversion[init.lvls.index(first)]
                        last = combo[3:]
                        lasts.append(last)
                        last = init.lvl_conversion[init.lvls.index(last)]
                        if first > last:  # Swap them if entered incorrectly
                            first = last
                            firsts[-1] = lasts[-1]
                            last = first
                            lasts[-1] = firsts[-1]
                        if last-1 >= len(init.lvls) or first < 0:
                            raise Exception("Combine Levels for", ev, "has a number outside 1-6")
                        for j in range(last - first):
                            lvlnames_c[first + j].clear()
                            concat_data_c.append(contestant_data['Couple'][first + j])
                            lvlcombos_c[first + j].clear()
                        concat_data_c.append(contestant_data['Couple'][last])
                        lvlcombos_c[last] = [pd.concat(concat_data_c)]
                    for i, entry in enumerate(lvlnames_c):
                        if entry != []:
                            if entry[0] in lasts:
                                tmp = entry[0]
                                entry[0] = firsts[lasts.index(entry[0])] + "-" + entry[0]
                                del firsts[lasts.index(tmp)]
                            eventlvlnames_c.append(entry[0])
                    for i, entry in enumerate(lvlcombos_c):
                        if entry != []:
                            eventlvls_c.append(entry[0])
                    contestant_data["Couple"] = eventlvls_c

                # Use Combine Levels field to set up dance dfs tree for Singles
                eventlvls_s = []
                eventlvlnames_s = []
                combinelvls_s = eventrow.loc[:, 'Combine Levels'][0]
                if type(combinelvls_s) == float:  # if blank
                    eventlvlnames_s = init.lvls.copy()
                else:
                    lvlnames_s = []
                    lvlcombos_s = []
                    firsts = []
                    lasts = []
                    for lvl in init.lvls:
                        lvlnames_s.append([lvl])
                    for dfs in contestant_data["Single"]:
                        lvlcombos_s.append([dfs])
                    for combo in combinelvl:
                        concat_data_s = []
                        first = combo[0:2]
                        firsts.append(first)
                        first = init.lvl_conversion[init.lvls.index(first)]
                        last = combo[3:]
                        lasts.append(last)
                        last = init.lvl_conversion[init.lvls.index(last)]
                        if first > last:  # Swap them if entered incorrectly
                            first = last
                            firsts[-1] = lasts[-1]
                            last = first
                            lasts[-1] = firsts[-1]
                        if last - 1 >= len(init.lvls) or first < 0:
                            raise Exception("Combine Levels for", ev, "has a number outside 1-6")
                        for j in range(last - first):
                            lvlnames_s[first + j].clear()
                            concat_data_s.append(contestant_data['Single'][first + j])
                            lvlcombos_s[first + j].clear()
                        concat_data_s.append(contestant_data['Single'][last])
                        lvlcombos_s[last] = [pd.concat(concat_data_s)]
                    for i, entry in enumerate(lvlnames_s):
                        if entry != []:
                            if entry[0] in lasts:
                                tmp = entry[0]
                                entry[0] = firsts[lasts.index(entry[0])] + "-" + entry[0]
                                del firsts[lasts.index(tmp)]
                            eventlvlnames_s.append(entry[0])
                    for i, entry in enumerate(lvlcombos_s):
                        if entry != []:
                            eventlvls_s.append(entry[0])
                    contestant_data["Single"] = eventlvls_s

                # Build Couples df for this event broken down by all metrics
                if ev in init.df_coup.columns:
                    shell_c = {}
                    for Couple, lvl in zip(contestant_data['Couple'], l_keys):
                        Couple = Couple[Couple[ev] > 0]
                        init.dance_dfs["C"][lvl] = {}
                        shell_c[lvl] = {}
                        if not Couple.empty:  # Couple df operations
                            Couple['type id'] = 'C'
                            Couple = Couple[['type id', 'Lead Dancer #', 'Lead First Name', 'Lead Last Name', 'Lead Age',
                                             'Follow Dancer #', 'Follow First Name', 'Follow Last Name', 'Follow Age',
                                             'Level', 'School', ev]]
                            if div.count('t') == 0 or div.count("T") == 0:
                                Couple["Instructor Dancer #'s"] = ''
                            for i, age in enumerate(eventages_c):
                                # Slice Couple so that it is inside age bracket
                                if i == 0:  # Set bounds of age bracket
                                    lower = 18
                                    upper = age
                                else:
                                    lower = eventages_c[i - 1] + 1
                                    upper = age
                                # Split the df based on which age is lower and then add them together at the end
                                sliced_f = Couple[Couple["Lead Age"] >= Couple["Follow Age"]]
                                sliced_f = sliced_f[(lower <= sliced_f["Follow Age"]) & (sliced_f["Follow Age"] <= upper)]
                                sliced_l = Couple[Couple["Lead Age"] < Couple["Follow Age"]]
                                sliced_l = sliced_l[(lower <= sliced_l["Lead Age"]) & (sliced_l["Lead Age"] <= upper)]
                                shell_c[lvl][age] = pd.concat([sliced_l, sliced_f])
                        else:
                            for i, age in enumerate(eventages_c):
                                init.dance_dfs["C"][lvl][age] = pd.DataFrame()
                                shell_c[lvl][age] = pd.DataFrame()

                    # Add together dfs based on the division metrics of this event
                    if div.count('A') == 0 and div.count('a') == 0:
                        # Couple Concat
                        for lvl in shell_c.keys():
                            for i, key in enumerate(shell_c[lvl].keys()):
                                if i == 0:
                                    tmp = shell_c[lvl][key]
                                    continue
                                tmp = pd.concat([tmp, shell_c[lvl][key]])
                            shell_c[lvl] = tmp

                    # Combine all levels if needed
                    if div.count('L') == 0 and div.count('l') == 0:
                        for i, key in enumerate(shell_c.keys()):
                            if i == 0:
                                tmp = shell_c[key]
                                continue
                            if type(tmp) is dict:
                                for subkey in tmp.keys():
                                    tmp[subkey] = pd.concat([tmp[subkey], shell_c[key][subkey]])
                            else:
                                tmp = pd.concat([tmp, shell_c[key]])
                        shell_c = tmp

                    # put shell into df
                    init.dance_dfs["C"] = shell_c
                else:
                    del init.dance_dfs["C"]

                if ev in init.df_sing.columns:
                    shell_s = {"Lead": {}, "Follow": {}}
                    shell_i = {"Lead": {}, "Follow": {}}
                    for Single, lvl in zip(contestant_data['Single'], l_keys):
                        Single = Single[Single[ev] > 0]
                        init.dance_dfs["S"]["Lead"][lvl] = {}
                        init.dance_dfs["S"]["Follow"][lvl] = {}
                        shell_s["Lead"][lvl] = {}
                        shell_i["Lead"][lvl] = {}
                        shell_s["Follow"][lvl] = {}
                        shell_i["Follow"][lvl] = {}
                        if not Single.empty:
                            if not Single[Single['Lead/Follow'] == 'Lead'].empty:
                                df = Single[(Single['Lead/Follow'] == 'Lead') | (Single['Lead/Follow'] == 'L')]
                                df.loc[:,'type id'] = 'L'
                                df = df.rename(columns={'First Name': 'Lead First Name', 'Last Name': 'Lead Last Name',
                                                        'Dancer #': 'Lead Dancer #', "Age": "Lead Age"})
                                df['Follow First Name'] = ''
                                df['Follow Last Name'] = ''
                                df['Follow Age'] = ''
                                df['Follow Dancer #'] = ''
                                df = df[['type id', 'Lead Dancer #', 'Lead First Name', 'Lead Last Name', "Lead Age", "Follow Age", "Instructor Dancer #'s",
                                         'Follow Dancer #', 'Follow First Name', 'Follow Last Name', 'Level', 'School', ev]]
                                for i, age in enumerate(eventages_s):
                                    # SLice so that it is inside age bracket
                                    if i == 0:  # Set bounds of age bracket
                                        lower = 18
                                        upper = age
                                    else:
                                        lower = eventages_s[i - 1] + 1
                                        upper = age
                                    # Split the df based on which age is lower and then add them together at the end
                                    sliced_l = df[(lower <= df["Lead Age"]) & (df["Lead Age"] <= upper)]
                                    shell_s["Lead"][lvl][age] = sliced_l
                                    shell_i["Lead"][lvl][age] = {}
                                    # Find unique instructors for this division
                                    for row, data in sliced_l.iterrows():  # Iterate down all rows of Single df
                                        if type(data["Instructor Dancer #'s"]) == int:
                                            tmp = [data["Instructor Dancer #'s"]]
                                        else:
                                            tmp = [int(x) for x in data["Instructor Dancer #'s"].split(";")]
                                        data["Instructor Dancer #'s"] = tmp
                                        del tmp
                                        for num in data["Instructor Dancer #'s"]:  # Iterate through all #'s in instructor lists
                                            if num in shell_i["Lead"][lvl][age].keys():
                                                shell_i["Lead"][lvl][age][num] += 1
                                            else:
                                                shell_i["Lead"][lvl][age][num] = 1
                            else:
                                for i, age in enumerate(eventages_s):
                                    init.dance_dfs["S"]["Lead"][lvl][age] = pd.DataFrame()
                                    shell_s["Lead"][lvl][age] = pd.DataFrame()
                                    shell_i["Lead"][lvl][age] = pd.DataFrame()

                            if not Single[Single['Lead/Follow'] == 'Follow'].empty:

                                df2 = Single[Single['Lead/Follow'] == 'Follow']
                                df2.loc[:,'type id'] = 'F'
                                df2 = df2.rename(columns={'First Name': 'Follow First Name', 'Last Name': 'Follow Last Name',
                                                    'Dancer #': 'Follow Dancer #', "Age": "Follow Age"})
                                df2['Lead First Name'] = ''
                                df2['Lead Last Name'] = ''
                                df2['Lead Age'] = ''
                                df2['Lead Dancer #'] = ''
                                df2 = df2[['type id', 'Lead Dancer #', 'Lead First Name', 'Lead Last Name', "Lead Age" , "Instructor Dancer #'s",
                                         'Follow Dancer #', 'Follow First Name', 'Follow Last Name', 'Follow Age', 'Level', 'School', ev]]
                                for i, age in enumerate(eventages_s):
                                    # SLice Couple so that it is inside age bracket
                                    if i == 0:  # Set bounds of age bracket
                                        lower = 18
                                        upper = age
                                    else:
                                        lower = eventages_s[i - 1] + 1
                                        upper = age
                                    # Split the df based on which age is lower and then add them together at the end
                                    sliced_f = df2[(lower <= df2["Follow Age"]) & (df2["Follow Age"] <= upper)]
                                    shell_s["Follow"][lvl][age] = sliced_f
                                    shell_i["Follow"][lvl][age] = {}

                            else:
                                for i, age in enumerate(eventages_s):
                                    init.dance_dfs["S"]["Follow"][lvl][age] = pd.DataFrame()
                                    shell_s["Follow"][lvl][age] = pd.DataFrame()
                                    shell_i["Follow"][lvl][age] = pd.DataFrame()
                        else:
                            for i, age in enumerate(eventages_s):
                                init.dance_dfs["S"]["Lead"][lvl][age] = pd.DataFrame()
                                init.dance_dfs["S"]["Follow"][lvl][age] = pd.DataFrame()
                                shell_s["Follow"][lvl][age] = pd.DataFrame()
                                shell_s["Lead"][lvl][age] = pd.DataFrame()
                                shell_i["Lead"][lvl][age] = pd.DataFrame()
                                shell_i["Follow"][lvl][age] = pd.DataFrame()

                        # Add together dfs based on the division metrics of this event
                        if div.count('A') == 0 and div.count('a') == 0:
                            # Lead concat
                            for lvl in shell_s["Lead"].keys():
                                for i, key in enumerate(shell_s["Lead"][lvl].keys()):
                                    if i == 0:
                                        tmp = shell_s["Lead"][lvl][key]
                                        continue
                                    tmp = pd.concat([tmp, shell_s["Lead"][lvl][key]])
                                shell_s["Lead"][lvl] = tmp

                            # Combine Instructor data
                            for lvl in shell_i["Lead"].keys():
                                for i, key in enumerate(shell_i["Lead"][lvl].keys()):
                                    if i == 0:
                                        tmp = shell_i["Lead"][lvl][key]
                                        continue
                                    tmp.update(shell_i["Lead"][lvl][key])
                                shell_i["Lead"][lvl] = tmp

                            # Follow Concat
                            for lvl in shell_s["Follow"].keys():
                                for i, key in enumerate(shell_s["Follow"][lvl].keys()):
                                    if i == 0:
                                        tmp = shell_s["Follow"][lvl][key]
                                        continue
                                    tmp = pd.concat([tmp, shell_s["Follow"][lvl][key]])
                                shell_s["Follow"][lvl] = tmp

                        # Combine all levels if needed
                        if div.count('L') == 0 and div.count('l') == 0:
                            # Lead Concat
                            for i, lvl in enumerate(shell_s["Lead"].keys()):
                                if i == 0:
                                    tmp = shell_s["Lead"][lvl]
                                    continue
                                if type(tmp) is dict:
                                    for subkey in tmp.keys():
                                        tmp[subkey] = pd.concat([tmp[subkey], shell_s["Lead"][lvl][subkey]])
                                else:
                                    tmp = pd.concat([tmp, shell_s["Lead"][lvl]])
                            shell_s["Lead"] = tmp

                            # Follow Concat
                            for i, lvl in enumerate(shell_s["Follow"].keys()):
                                if i == 0:
                                    tmp = shell_s["Follow"][lvl]
                                    continue
                                if type(tmp) is dict:
                                    for subkey in tmp.keys():
                                        tmp[subkey] = pd.concat([tmp[subkey], shell_s["Follow"][lvl][subkey]])
                                else:
                                    tmp = pd.concat([tmp, shell_s["Follow"][lvl]])
                            shell_s["Follow"] = tmp

                        # if event Singles should be combined
                        if (div.count('S') == 0 and div.count('s') == 0) or (div.count('T') == 0 and div.count('t') == 0):
                            # Singles Lead/Follow Concat
                            for i, key in enumerate(shell_s.keys()):
                                if i == 0:
                                    tmp = shell_s[key]
                                    continue
                                if type(tmp) is dict:  # Go down the tree and combine the corresponding N nodes
                                    for subkey in tmp.keys():
                                        if type(tmp[subkey]) is dict:
                                            for subkey2 in tmp[subkey].keys():
                                                tmp[subkey][subkey2] = pd.concat(
                                                    [tmp[subkey][subkey2], shell_s[key][subkey][subkey2]])
                                        else:
                                            tmp[subkey] = pd.concat([tmp[subkey], shell_s[key][subkey]])
                                else:
                                    tmp = pd.concat([tmp, shell_s[key]])
                            shell_s = tmp

                        # put shells into df
                        init.dance_dfs["S"] = shell_s
                else:
                    del init.dance_dfs["S"]
                # # Add together dfs based on the division metrics of this event
                # if div.count('A') == 0 and div.count('a') == 0:
                #     # Lead concat
                #     for lvl in shell_s["Lead"].keys():
                #         for i, key in enumerate(shell_s["Lead"][lvl].keys()):
                #             if i == 0:
                #                 tmp = shell_s["Lead"][lvl][key]
                #                 continue
                #             tmp = pd.concat([tmp, shell_s["Lead"][lvl][key]])
                #         shell_s["Lead"][lvl] = tmp
                #     # Combine Instructor data
                #     for lvl in shell_i["Lead"].keys():
                #         for i, key in enumerate(shell_i["Lead"][lvl].keys()):
                #             if i == 0:
                #                 tmp = shell_i["Lead"][lvl][key]
                #                 continue
                #             tmp.update(shell_i["Lead"][lvl][key])
                #         shell_i["Lead"][lvl] = tmp
                #
                #
                #     # Follow Concat
                #     for lvl in shell_s["Follow"].keys():
                #         for i, key in enumerate(shell_s["Follow"][lvl].keys()):
                #             if i == 0:
                #                 tmp = shell_s["Follow"][lvl][key]
                #                 continue
                #             tmp = pd.concat([tmp, shell_s["Follow"][lvl][key]])
                #         shell_s["Follow"][lvl] = tmp
                #
                #     # Couple Concat
                #     for lvl in shell_c.keys():
                #         for i, key in enumerate(shell_c[lvl].keys()):
                #             if i == 0:
                #                 tmp = shell_c[lvl][key]
                #                 continue
                #             tmp = pd.concat([tmp, shell_c[lvl][key]])
                #         shell_c[lvl] = tmp
                #
                # # Combine all levels if needed
                # if div.count('L') == 0 and div.count('l') == 0:
                #     # Lead Concat
                #     for i, lvl in enumerate(shell_s["Lead"].keys()):
                #         if i == 0:
                #             tmp = shell_s["Lead"][lvl]
                #             continue
                #         if type(tmp) is dict:
                #             for subkey in tmp.keys():
                #                 tmp[subkey] = pd.concat([tmp[subkey], shell_s["Lead"][lvl][subkey]])
                #         else:
                #             tmp = pd.concat([tmp, shell_s["Lead"][lvl]])
                #     shell_s["Lead"] = tmp
                #
                #     # Follow Concat
                #     for i, lvl in enumerate(shell_s["Follow"].keys()):
                #         if i == 0:
                #             tmp = shell_s["Follow"][lvl]
                #             continue
                #         if type(tmp) is dict:
                #             for subkey in tmp.keys():
                #                 tmp[subkey] = pd.concat([tmp[subkey], shell_s["Follow"][lvl][subkey]])
                #         else:
                #             tmp = pd.concat([tmp, shell_s["Follow"][lvl]])
                #     shell_s["Follow"] = tmp
                #
                #     for i, key in enumerate(shell_c.keys()):
                #         if i == 0:
                #             tmp = shell_c[key]
                #             continue
                #         if type(tmp) is dict:
                #             for subkey in tmp.keys():
                #                 tmp[subkey] = pd.concat([tmp[subkey], shell_c[key][subkey]])
                #         else:
                #             tmp = pd.concat([tmp, shell_c[key]])
                #     shell_c = tmp
                #
                # # if event Singles should be combined
                # if (div.count('S') == 0 and div.count('s') == 0) or (div.count('T') == 0 and div.count('t') == 0):
                #     # Singles Lead/Follow Concat
                #     for i, key in enumerate(shell_s.keys()):
                #         if i == 0:
                #             tmp = shell_s[key]
                #             continue
                #         if type(tmp) is dict: # Go down the tree and combine the corresponding N nodes
                #             for subkey in tmp.keys():
                #                 if type(tmp[subkey]) is dict:
                #                     for subkey2 in tmp[subkey].keys():
                #                         tmp[subkey][subkey2] = pd.concat([tmp[subkey][subkey2], shell_s[key][subkey][subkey2]])
                #                 else:
                #                     tmp[subkey] = pd.concat([tmp[subkey], shell_s[key][subkey]])
                #         else:
                #             tmp = pd.concat([tmp, shell_s[key]])
                #     shell_s = tmp
                #
                # # put shells into df
                # init.dance_dfs["S"] = shell_s
                # init.dance_dfs["C"] = shell_c

                # If Couples and Singles need to be combined
                if div.count('T') == 0 and div.count('t') == 0 and (init.dance_dfs.get("S") is not None) and (init.dance_dfs.get("C") is not None):
                    for i, key in enumerate(init.dance_dfs.keys()):
                        if i == 0:
                            tmp = init.dance_dfs[key]
                            continue
                        if type(tmp) is dict:  # Go down the tree and combine the corresponding N nodes
                            for subkey in tmp.keys():
                                if type(tmp[subkey]) is dict:
                                    for subkey2 in tmp[subkey].keys():
                                        tmp[subkey][subkey2] = pd.concat([tmp[subkey][subkey2], init.dance_dfs[key][subkey][subkey2]])
                                else:
                                    tmp[subkey] = pd.concat([tmp[subkey], init.dance_dfs[key][subkey]])
                        else:
                            tmp = pd.concat([tmp, init.dance_dfs[key]])
                    init.dance_dfs["A"] = tmp
                    del init.dance_dfs["S"]
                    del init.dance_dfs["C"]

                    del shell_s
                    del shell_c

                # df = pd.DataFrame()
                # # Single df operations both for lead and follow
                # if not Single.empty:
                #     # Find totals for this Singles level
                #     total_contestants_s = total_contestants_s + Single.shape[0]
                #     total_entries_s = total_entries_s + Single[ev].sum()
                #     # Track Levels largest to smallest
                #     added = False
                #     for i, lev in enumerate(decend_level_s):
                #         if lev[1] <= Single.shape[0]:
                #             added = True
                #             decend_level_s.insert(i-1, [lvl, Single.shape[0]])
                #             # decend_level_s.append([lvl, Single.shape[0]]) for testing
                #             break
                #     if not added:
                #         decend_level_s.append([lvl, Single.shape[0]])

                # # Set up Percent trackers
                # for key in keys:
                #     if init.dance_dfs_dict_s[key].empty:
                #         lev_percent_s.append(-1)
                #         percent_left_s.append(-1)
                #     else:
                #         lev_percent_s.append(init.dance_dfs_dict_s[key][ev].sum() / total_entries_s)
                #         percent_left_s.append(1.0)
                #
                #     if init.dance_dfs_dict_c[key].empty:
                #         lev_percent_c.append(-1)
                #         percent_left_c.append(-1)
                #     else:
                #         lev_percent_c.append(init.dance_dfs_dict_c[key][ev].sum() / total_entries_c)
                #         percent_left_c.append(1.0)

                # # Set Holes
                # for key in keys:
                #     if init.dance_dfs_dict_s[key].empty:
                #         tot_holes.append(-1)
                #     else:
                #         tot_holes.append(0)

                # Check if current dfs are empty, and remove them
                deleteEmpty(init.dance_dfs)
                # if both Pools are totally empty continue
                # In this case it would only happen if there were no entries to current dance 'ev'
                floors = eventrow.loc[:, 'Floors'][0]
                heat_list = HeatList([], floors, couples_per_floor, eventages_s, eventages_c, eventlvlnames_s, eventlvlnames_c)  # list of individual heats for the current dance 'ev'
                if len(init.dance_dfs) == 0:
                    print("No data in this category", ev)
                    heat_list = -1
                    continue
                pre_dance_dfs = init.dance_dfs.copy()
                init.inst_tree = buildInstTree(init.dance_dfs, {}, ev)
                init.inst2sing_tree = buildInst2SingTree(init.dance_dfs, {}, ev)
                pre_inst_tree = init.inst_tree.copy()
                del init.dance_dfs["S"]
# ---------------------------------------------- Selection Process -----------------------------------------------------
                if init.dance_dfs.get("S") is None:
                    singles_empty = True
                else:
                    singles_empty = False  # True when all singles are in heats for this event 'ev'
                if init.dance_dfs.get("C") is None:
                    couples_empty = True
                else:
                    couples_empty = False  # True when all singles are in heats for this event 'ev'
                split_mode = False
                while (init.dance_dfs.get("D") is None):  # while there are contestants in the dfs still
                    heat_roster = []  # holds full contestant data for a heat, each element is a list for a dance floor
                    fin_rooms = []  # marks a room as finished with a 1
                    instructors_in_heat = []  # Holds instructor numbers for quick reference
                    singles_in_heat = []  # Holds contestant number for quick reference
                    couples_in_heat = []  # Holds Couples number for quick ref.
                    # couples_in_heat format [[1 Lead #,1 Follow #,...,N Follow #],...[1 Lead #,1 Follow #,..., N Follow #]]
                    floor_info = pickDfs(ev, init.dance_dfs, init.inst_tree, floors, div, eventages_s, eventages_c, couples_per_floor)
                    if len(floor_info) < floors:
                        split_mode = True
                        vacant_rooms = floors - len(floor_info)
                    for floor in range(floors):
                        heat_roster.append([])
                        instructors_in_heat.append([])
                        singles_in_heat.append([])
                        couples_in_heat.append([])
                        fin_rooms.append(0)
                    heat_finished = False
                    # selection_dfs = []
                    s_floors = []
                    sfin_rooms = []
                    c_floors = []
                    cfin_rooms = []
                    # Separate out singles and couple assigned floors
                    if 't' in div or 'T' in div:
                        for floor in floor_info:
                            if floor[0] == "S":
                                s_floors.append(floor)
                                sfin_rooms.append(0)
                            else:
                                c_floors.append(floor)
                                cfin_rooms.append(0)
                    heat_key = each + '-' + every + '-' + ev + '-' + str(len(heat_list.getRostersList()))
                    log = ConflictLog(floor_info)  # make conflict log for this heat
                    heat = Heat(heat_key, floor_info, heat_roster, singles_in_heat, instructors_in_heat, couples_in_heat)
                    while not heat_finished:
                        # ---------------------------------------------- Singles ---------------------------------------
                        heat_holes = []
                        for place in floor_info:
                            heat_holes.append(0)
                        # Make Instructors for heat list that will change based on placed contestants and instructors
                        dance_df = []
                        inst_tree_nodes = []
                        inst2sing_tree_nodes = []
                        instructors_available_for_heat = []
                        for roomid, info in enumerate(s_floors):
                            for i, key in enumerate(info):
                                if i == 0:
                                    tmp = init.dance_dfs[key]
                                    tmp2 = init.inst_tree[key]
                                    tmp3 = init.inst2sing_tree[key]
                                elif type(tmp) is dict:
                                    try:
                                        tmp = tmp[key]
                                        tmp2 = tmp2[key]
                                        tmp3 = tmp3[key]
                                    except:
                                        print(info, "Division removed already, likely due to split where this division is empty")
                                        tmp = pd.DataFrame()
                                        tmp2 = pd.DataFrame()
                                        tmp3 = pd.DataFrame()
                            dance_df.append(tmp)
                            inst_tree_nodes.append(tmp2)
                            inst2sing_tree_nodes.append(tmp3)
                            instructors_available_for_heat.append(list(tmp2.keys()))
                        selection_finished = False
                        print("Event " + ev + ", Heat number: " + str(heat_list.getHeatCount()+1))
                        print(floor_info)
                        print()
                        while not selection_finished:
                            if len(s_floors) == 0:  # If no singles move to couples selection
                                break
                            for roomid, room_info in enumerate(s_floors):  # For each ballroom make 1 instructor/single pair
                                print('Selecting for room', roomid, room_info)
                                if sfin_rooms[roomid] > 0:  # if room is filled or declared full, continue on
                                    continue
                                placed = False  # Set when a valid instructor/single pair is placed
                                failed = False
                                consecutive = 0  # Stops infinite loops on failed attempts to add candidate to the heat
                                solved = 0  # counts how many resolveConflict() calls for this pair search
                                while (not placed) and not (sfin_rooms[roomid] > 0):  # Find a viable single/inst match
                                    if consecutive > max_conflicts:
                                        # first update the dance dfs
                                        for df, info in zip(dance_df, s_floors):
                                            updateDanceDfs(init.dance_dfs, df, info)
                                        resolve = resolveConflictSingles(roomid, dance_df, inst_tree_nodes, log, heat, heat_list, solved, instructors_available_for_heat, inst2sing_tree_nodes, ev)
                                        # TODO test the meta data lists after a resolve
                                        if resolve == 1:
                                            failed = False
                                            consecutive = 0
                                            solved += 1
                                        elif not failed: #and len(instructors_available_for_heat)[roomid] < len(list(inst_tree_nodes[roomid].keys())):
                                            # If first time failing try to add the instructors in the pool back into the picking list, the selection earlier might have just been unlucky in selection
                                            # meaning it saved a subset of similar instructor clusters stopping the selection process from even considering them
                                            failed = True
                                            print("Failed to resolve conflict adding back instructors in pool for more selector options")
                                            instructors_available_for_heat[roomid] = list(inst_tree_nodes[roomid].keys())
                                        elif failed:
                                            sfin_rooms[roomid] = 2  # force a finish
                                            if (sfin_rooms.count(1) + sfin_rooms.count(2)) == len(s_floors):
                                                selection_finished = True
                                            continue
                                    inst = random.choice(instructors_available_for_heat[roomid])  # get random instructor, will throw error if list at index is empty
                                    instructor_taken = False
                                    for selection in instructors_in_heat:
                                        if instructor_taken:
                                            break
                                        if selection.count(inst) > 0:  # Check if instructor is being used in heat
                                            instructor_taken = True
                                            break
                                    if instructor_taken:
                                        # cons = getContestantConflictList(dance_df[roomid], inst, singles_in_heat)
                                        log.addConflict(ConflictItemSingle(1, inst), roomid)
                                        consecutive += 1
                                        continue
                                # Find a single to pair with this instructor
                                    found = False
                                    # Loop over each row in the df for this level
                                    try:
                                        df_shuffled = dance_df[roomid].sample(frac=1)
                                    except:
                                        pass
                                    for row, entry in df_shuffled.iterrows():
                                        next_contestant = False
                                        if found:
                                            break
                                        # Set column variables based on single type
                                        if entry["type id"] == "L":
                                            contestant_col = "Lead Dancer #"
                                            inst_col = "Follow Dancer #"
                                            inst_fname = "Follow First Name"
                                            inst_lname = "Follow Last Name"
                                        elif entry["type id"] == "F":
                                            contestant_col = "Follow Dancer #"
                                            inst_col = "Lead Dancer #"
                                            inst_fname = "Lead First Name"
                                            inst_lname = "Lead Last Name"
                                        else:
                                            raise Exception("Type id for " + entry + " is invalid")
                                        # Retrofit the instructor list entry
                                        if type(entry["Instructor Dancer #'s"]) == int:
                                            tmp = [entry["Instructor Dancer #'s"]]
                                        else:
                                            tmp = [int(x) for x in entry["Instructor Dancer #'s"].split(";")]
                                        entry["Instructor Dancer #'s"] = tmp
                                        del tmp
                                        # Loop over the instructor list for the contestant row
                                        for num in entry["Instructor Dancer #'s"]:
                                            if num == inst:
                                                for i, rost in enumerate(singles_in_heat):
                                                    if rost.count(entry[contestant_col]) > 0:
                                                        next_contestant = True
                                                if not next_contestant:
                                                    # Set candidate to this single/inst match
                                                    candidate = entry.to_frame().T
                                                    candidate = candidate.reset_index(drop=True)
                                                    instructor_data = init.df_inst[init.df_inst["Dancer #"] == inst].reset_index(drop=True)
                                                    candidate.loc[0, inst_col] = instructor_data.loc[0, "Dancer #"]
                                                    candidate.loc[0, inst_fname] = instructor_data.loc[0, "First Name"]
                                                    candidate.loc[0, inst_lname] = instructor_data.loc[0, "Last Name"]
                                                    found = True
                                                    break
                                    if found:
                                        # Remove/add to tracking data structs
                                        heat.addEntry(candidate, roomid)
                                        heat_roster = heat.getRoster()
                                        instructors_in_heat = heat.getInstructors()
                                        singles_in_heat = heat.getSingles()
                                        instructors_available_for_heat[roomid].remove(inst)
                                        print("Candidate placed: ")
                                        print(candidate.loc[0, contestant_col], inst, 'room', roomid)
                                        print()
                                        for num in candidate.loc[:, "Instructor Dancer #'s"][0]:
                                            if inst_tree_nodes[roomid][num] == 1:
                                                del inst_tree_nodes[roomid][num]
                                                if num in instructors_available_for_heat[roomid]:
                                                    instructors_available_for_heat[roomid].remove(num)
                                            else:
                                                inst_tree_nodes[roomid][num] -= 1
                                        # Remove the placed candidate from the df, or -1 if multi-entry
                                        if candidate.loc[:, ev][0] == 1:
                                            dance_df[roomid] = dance_df[roomid].drop(dance_df[roomid][dance_df[roomid][contestant_col] == candidate.loc[0, contestant_col]].index)
                                        else:
                                            dance_df[roomid].loc[dance_df[roomid].loc[:, contestant_col] == candidate.loc[0, contestant_col], ev] = candidate.loc[0, ev] - 1
                                        placed = True
                                        # Check if current level df is empty after a placed candidate
                                        if dance_df[roomid].empty:
                                            sfin_rooms[roomid] = 1
                                            updateDanceDfs(init.dance_dfs, dance_df[roomid], room_info)
                                            deleteEmpty(init.dance_dfs)
                                            if init.dance_dfs.get("S") is None:
                                                singles_empty = True
                                            if heat.getRoster()[roomid] < acceptablecouples:
                                                PoachPrevHeatsSingles(roomid, room_info, dance_df[roomid], heat, heat_list, acceptablecouples)
                                    else:
                                        log.addConflict(ConflictItemSingle(2, inst), roomid)
                                        consecutive += 1
                                    # Determine if room is finished
                                    if len(heat_roster[roomid]) == couples_per_floor:
                                        sfin_rooms[roomid] = 1
                                        if dance_df[roomid].shape[0] < int(couples_per_floor / 2) and dance_df[roomid].shape[0] != 0:
                                            if dance_df[roomid][ev].sum() < heat_list.getDivisionHeatCount(room_info) * 2:
                                                # backfill(dance_df[roomid], room_info, heat_list, couples_per_ballroom, ev, init.df_inst)
                                                updateDanceDfs(init.dance_dfs, dance_df[roomid], room_info)
                                                deleteEmpty(init.dance_dfs)
                                                if init.dance_dfs.get("S") is None:
                                                    singles_empty = True
                                            else:
                                                print("Not enough heats to backfill")
                                    if singles_empty:
                                        print("Singles Empty for", ev)
                                        for i, info in enumerate(sfin_rooms):
                                            if sfin_rooms[i] == 0:
                                                sfin_rooms[i] = 1
                                    if len(instructors_available_for_heat[roomid]) == 0:
                                        sfin_rooms[roomid] = 1
                                    if solved == 1000 and len(heat_roster[roomid]) != couples_per_floor:
                                        sfin_rooms[roomid] = 2  # Forced finish
                                    if consecutive > max_conflicts and (heat_list.getDivisionHeatCount(room_info) == 0): # if conflicts and no previous heats
                                        print("Initial Heat for ", room_info, " forced finish")
                                        print("df size:", dance_df[roomid].shape[0])
                                        sfin_rooms[roomid] = 2  # Forced finish
                                    # Determine if heat finished
                                    if sfin_rooms[roomid] > 0 and (len(heat.getSingles()[roomid]) < acceptablecouples):
                                        PoachPrevHeatsSingles(roomid, room_info, dance_df[roomid], heat, heat_list, acceptablecouples)
                                    if (sfin_rooms.count(1) + sfin_rooms.count(2)) == len(s_floors):
                                        selection_finished = True

                        # selection_dfs.append(dance_df[:])
                            # Selection finished
                            # # Update Holes metadata
                            # if not lev_mode_selection:
                            #     pass
                            #     # # Add number of holes before a level mode selection
                            #     # for i, lvl in enumerate(heat_room_levs):
                            #     #     tot_holes[lvl] += couples_per_ballroom - len(heat_roster[i])
                            #     #     heat_holes[i] = couples_per_ballroom - len(heat_roster[i])
                            # else:
                            #     # Add holes after level mode selection but only to newly filled room
                            #     i = vacant_rooms  # Start of the vacant rooms
                            #     for lvl in level_mode_selection_lev:
                            #         tot_holes[lvl] += (couples_per_ballroom - len(heat_roster[i]))
                            #         heat_holes[i] = (couples_per_ballroom - len(heat_roster[i]))
                            #         i += 1

                        # Heat fully complete Construct key for the new heat
                        # Fill the unused room data if needed
                        # if len(heat_room_levs) < ballrooms:
                        #     for i in range(ballrooms-len(heat_room_levs)):
                        #         heat_room_levs.append(-1)
                        #         heat_holes.append(couples_per_ballroom)
                        # Update Dance dfs with all placements
                        for df, info in zip(dance_df, s_floors):
                            updateDanceDfs(init.dance_dfs, df, info)
                        deleteEmpty(init.dance_dfs)  # Remove the keys from the tree that are empty
                        # ---------------------------------------------- Couples ---------------------------------------
                        if len(c_floors) > 0:  # if a couples key has been picked
                            print("Selecting Couples")
                            for roomid, info in enumerate(c_floors):
                                for i, key in enumerate(info):
                                    if i == 0:
                                        tmp = init.dance_dfs[key]
                                        continue
                                    if type(tmp) is dict:
                                        try:
                                            tmp = tmp[key]
                                        except:
                                            print(info, "Division removed already, likely due to split where this division is empty")
                                            tmp = pd.DataFrame()
                                dance_df.append(tmp)
                            # Loop rooms, backfill or filling new room to completion before moving to the next room
                            for roomid, room_info in enumerate(c_floors):
                                # If room finished
                                if cfin_rooms[roomid] > 0:
                                    continue
                                print('Selecting for room', roomid, room_info)
                                consecutive = 0  # Stops infinite looping continual failed attempts to add candidate to the heat
                                solved = 0  # counts how many resolveConflict() calls for this heat
                                # Get suitable candidates from the df of the current division
                                if len(s_floors) > 0:
                                    singles_index = len(s_floors)-1
                                else:
                                    singles_index = 0
                                while len(heat_roster[roomid+singles_index]) < max_dance_couples and cfin_rooms[roomid] == 0:
                                    if consecutive > max_conflicts:
                                        resolve = resolveConflictCouples(roomid, dance_df, inst_tree_nodes, log, heat, heat_list, solved, instructors_available_for_heat, inst2sing_tree_nodes, ev)
                                        # TODO CHeck the meta data lists after a resolve
                                        if resolve == 1:
                                            consecutive = 0
                                        else:
                                            cfin_rooms[roomid] = 2  # force a finish
                                            if (cfin_rooms.count(1) + cfin_rooms.count(2)) == len(c_floors):
                                                break
                                            continue
                                    candidate = dance_df[roomid+singles_index].sample(ignore_index=True)  # Pick out random entry from df, may need to have try catch
                                    # Check candidate has no one already in heat
                                    dup_sing = False
                                    dup_inst = False
                                    dup_coup = False
                                    if candidate.loc[0, 'type id'] == 'C':
                                        number = candidate.loc[0, 'Lead Dancer #']
                                        fnumber = candidate.loc[0, 'Follow Dancer #']
                                        # TODO copy this section to 'A' as well
                                        # Check contestants in heat
                                        for selection in singles_in_heat:
                                            if dup_sing:
                                                break
                                            if selection.count(number) > 0:  # Check if Lead is being used in heat
                                                dup_sing = True
                                                break
                                            if selection.count(fnumber) > 0:  # Check if Follow is being used in heat
                                                dup_sing = True
                                                break
                                        # Check Instructors in heat
                                        for selection in instructors_in_heat:
                                            if dup_inst:
                                                break
                                            if selection.count(number) > 0:  # Check if Lead is being used in heat
                                                dup_inst = True
                                                break
                                            if selection.count(fnumber) > 0:  # Check if Follow is being used in heat
                                                dup_inst = True
                                                break
                                        # Check couples in heat
                                        for selection in couples_in_heat:
                                            if dup_coup:
                                                break
                                            if selection.count(number) > 0:  # Check if Lead is being used in heat
                                                dup_coup = True
                                                break
                                            if selection.count(fnumber) > 0:  # Check if Follow is being used in heat
                                                dup_coup = True
                                                break
                                    if dup_sing or dup_inst or dup_coup:
                                        consecutive += 1  # stop infinite looping when list has no candidate to add to this heat
                                        # Add to the conflict log
                                        if dup_sing:
                                            conflict = ConflictItemCouple(2, number, fnumber)
                                            log.addConflict(conflict, roomid+singles_index)
                                            continue
                                        if dup_coup:
                                            conflict = ConflictItemCouple(1, number, fnumber)
                                            log.addConflict(conflict, roomid+singles_index)
                                            continue
                                        if dup_inst:
                                            raise Exception("Couple in Instructor list", number, fnumber)
                                    # # Check if instructor inside heat, only for Singles
                                    # if candidate.loc[:, 'type id'][0] != "C":  # if not a couple entry
                                    #     # Set which column will be used based on Leader or Follower single
                                    #     if candidate.loc[:, 'type id'][0] == 'F':
                                    #         inst_col = 'Lead Dancer #'
                                    #     else:
                                    #         inst_col = 'Follow Dancer #'
                                    #
                                    #     potential_instructors = candidate[inst_col].tolist()[0]
                                    #     # loop through potential_instructors list,
                                    #     # if not in heat already, keep the list in case of a later swap
                                    #     # add to heat roster and instructors_in_heat
                                    #     for inst in potential_instructors:
                                    #         found = False
                                    #
                                    #         for lev in instructors_in_heat:
                                    #         for lev in instructors_in_heat:
                                    #             if lev.count(inst) != 0:
                                    #                 found = True
                                    #                 break
                                    #         if not found:
                                    #             instructors_in_heat[roomlvl].append(inst)
                                    #             added = True
                                    #             break
                                    # instructors_in_heat[roomlvl].append(-1)
                                    else:  # Else = there are no conflicts
                                        # if candidate possible add to the roster, remove that entry from the pool
                                        heat.addEntry(candidate, roomid+singles_index)
                                        print("Candidate placed: ")
                                        print("Lead", number, "Follow", fnumber, "heat num", heat_list.getHeatCount()+1, 'room', roomid+singles_index, room_info)
                                        print()
                                        col = "Lead Dancer #"
                                        fcol = "Follow Dancer #"
                                        # if last or only entry remove it from query_df
                                        if candidate.loc[0, ev] == 1:
                                            dance_df[roomid+singles_index] = dance_df[roomid+singles_index].drop(dance_df[roomid+singles_index][(dance_df[roomid+singles_index][col] == candidate.loc[0, col]) & (dance_df[roomid+singles_index][fcol] == candidate.loc[0, fcol])].index)
                                        else:
                                            dance_df[roomid+singles_index].loc[(dance_df[roomid+singles_index].loc[:, col] == candidate.loc[0, col]) & (dance_df[roomid+singles_index].loc[:, fcol] == candidate.loc[0, fcol]), ev] = candidate.loc[0, ev] - 1
                                        # Add candidate data to tracker lists
                                        heat_roster = heat.getRoster()
                                        couples_in_heat = heat.getCouples()
                                        # Check if current df is empty after adding the candidate
                                        if dance_df[roomid+singles_index].empty:
                                            cfin_rooms[roomid] = 1
                                            updateDanceDfs(init.dance_dfs, dance_df[roomid+singles_index], room_info)
                                            deleteEmpty(init.dance_dfs)  # Clean up parent levels if needed
                                            if init.dance_dfs.get("C") is None:
                                                couples_empty = True
                                            if len(heat.getRoster()[roomid]) < acceptablecouples:
                                                PoachPrevHeatsCouples(roomid, room_info, dance_df[roomid], heat, heat_list, acceptablecouples)
                                    if len(heat_roster[roomid+singles_index]) == couples_per_floor:
                                        cfin_rooms[roomid] = 1
                                        if dance_df[roomid+singles_index].shape[0] < int(couples_per_floor/2) and dance_df[roomid+singles_index].shape[0] != 0:
                                            if dance_df[roomid+singles_index][ev].sum() < heat_list.getDivisionHeatCount(room_info)*2:
                                                # backfill(dance_df[roomid+singles_index], room_info, heat_list, couples_per_ballroom, ev, init.df_inst)
                                                updateDanceDfs(init.dance_dfs, dance_df[roomid+singles_index], room_info)
                                                deleteEmpty(init.dance_dfs)
                                                if init.dance_dfs.get("C") is None:
                                                    couples_empty = True
                                            else:
                                                print("Not enough heats to backfill")
                                        break
                                    if couples_empty:
                                        for i, info in enumerate(cfin_rooms):
                                            if cfin_rooms[i] == 0:
                                                cfin_rooms[i] = 1
                                    if solved == 1000 and len(heat_roster[roomid+singles_index]) != couples_per_floor:
                                        cfin_rooms[roomid] = 2  # Forced finish
                                    # if consecutive == max_conflicts and (heat_list.getDivisionHeatCount(room_info) == 0):  # if conflicts and no previous heats
                                    #     cfin_rooms[roomid] = 2  # Forced finish
                                    # Determine if heat finished
                                if (cfin_rooms.count(1) + cfin_rooms.count(2)) == len(c_floors):
                                    break
                            for df, info in zip(dance_df, c_floors):
                                updateDanceDfs(init.dance_dfs, df, info)
    # ------------------------------------------------- Double up on unused rooms --------------------------------------
                        # If levels < ballrooms and not in split mode already, figure out if the current selection pool can be put into open room(s)
                        if split_mode and (not (singles_empty and couples_empty)):
                            print('Spliting a Division on the floor')
                            # Sanity Check all the current floors
                            splits_count = [[0, 0, 0]]
                            splits_info = [[0]]
                            for i, info in enumerate(floor_info):
                                # Find meta data for singles list
                                if info[0] == "S" and init.dance_dfs.get("S") is not None:
                                    unique = findUnique(pre_inst_tree, s_floors, [])
                                    if not unique > couples_per_floor*(floor_info.count(info)+1):  # If room cannot be split due to instructor constraints
                                        continue
                                    count = findContestantCount(init.dance_dfs, info, ev)
                                    # if len(leftover_inst) < count:
                                    #     count = len(leftover_inst)
                                elif init.dance_dfs.get("C") is not None:
                                    count = findContestantCount(init.dance_dfs, info, ev)
                                if count == None:
                                    continue
                                count.append(int(count[0]/couples_per_floor))  # find out how many rooms div can occupy
                                for i, split in enumerate(splits_count):  # order largest to smallest count
                                    if split[1] > count[1]:
                                        continue
                                    else:
                                        splits_count.insert(i, count)
                                        splits_info.insert(i, info)
                                        break
                            # Loop over all rooms in heat
                            splits_count = splits_count[:-1]   # remove termination character
                            splits_info = splits_info[:-1]   # remove termination character
                            takefrom_count = []
                            for data in splits_count:
                                takefrom_count.append(data.copy())
                            takefrom_info = []
                            for data in splits_info:
                                takefrom_info.append(data.copy())
                            while split_mode:  # While loop to allow for multiple div split
                                splititer = 0
                                for i, vacant in enumerate(heat_roster):
                                    if len(vacant) > 0 or (len(floor_info)-1) >= i:  # If room is already in use, continue loop
                                        continue
                                    if len(takefrom_count) == 0:
                                        break
                                    # Choose the largest div, then re-organize the list after picking
                                    floor_info.append(takefrom_info[0])
                                    log.addRoom(takefrom_info[0])
                                    if 't' in div or 'T' in div:
                                        if splits_info[0][0] == "S":
                                            s_floors.append(takefrom_info[0])
                                            sfin_rooms.append(0)
                                        else:
                                            c_floors.append(takefrom_info[0])
                                            cfin_rooms.append(0)
                                    # Decrement the counts
                                    takefrom_count[0][0] -= couples_per_floor
                                    takefrom_count[0][1] -= couples_per_floor
                                    splits_count[splititer][0] -= couples_per_floor
                                    splits_count[splititer][1] -= couples_per_floor
                                    if takefrom_count[0][1] < couples_per_floor: # If next split cannot make a full room
                                        # If empty after selection or other splits exist
                                        if len(takefrom_count) > 1 or takefrom_count[0][1] <= 0:
                                            # replace = takefrom_count[0].copy()
                                            del takefrom_count[0]
                                            if splits_count[splititer][0] <= 0:
                                                del splits_count[splititer]
                                            splititer += 1
                                            del takefrom_info[0]
                                    # for i, split in enumerate(takefrom_count):  # order largest to smallest count
                                    #     if split[1][1] > replace[1]:
                                    #         continue
                                    #     else:
                                    #         splits_count.insert(i, count)
                                    #         splits_info.insert(i, info)
                                    #         break
                                if len(splits_count) == 0 or len(floor_info) == floors:
                                    split_mode = False
                            print("Done spliting new floor breakdown", floor_info)
                        else:
                            heat_finished = True
                    # add division element for empty rooms
                    while len(heat.getDiv()) < floors:
                        heat.getDiv().append([])
                    checkheat(heat)
                    heat_list.appendList(heat)  # add completed heat to the HeatList obj
                    split_mode = False
                    dance_heat_count += 1
                    tot_current_heats += 1
                    if tot_current_heats > max_heats:
                        print("Exceeded max heats for event time metrics")
                    if singles_empty and couples_empty:
                        init.dance_dfs["D"] = 1
                if heats[each][every].get(ev) is not None:
                    heats[each][every][ev] = heat_list
                    print(ev, "finished Selection with", dance_heat_count, "Heats")
    buildEvent(heats, eventName)


def makeHeatDict(genrelist, df_cat):
    heats = {}
    syllabus = ['Open', 'Closed']
    # For each genre, fill a dictionary key'd with genre, with a sub dict key'd by dance name with that genre
    # This will be used to hold all heats for that dance
    for each in genrelist:
        heats[each] = {}
        for every in syllabus:
            df_genre = df_cat[(df_cat.Genre == each) & (df_cat.Syllabus == every)]
            heats[each][every] = {}
            while not df_genre.empty:
                subdict = heats[each][every]
                subdict.update({df_genre["Dance"].iloc[0]: []})
                #heats.update(subdict.update({df_genre["Dance"].iloc[0]: []}))
                df_genre = df_genre.iloc[1:]
    return heats


def sliceDfs(df_sing, df_coup):
    # Slice Contestant dfs based on levels
    df_sing_AB = df_sing[(df_sing.Level == 'NC') | (df_sing.Level == 'B1') | (df_sing.Level == 'B2')]
    df_sing_FB = df_sing[(df_sing.Level == 'B3') | (df_sing.Level == 'B4')]
    df_sing_AS = df_sing[(df_sing.Level == 'S1') | (df_sing.Level == 'S2')]
    df_sing_FS = df_sing[(df_sing.Level == 'S3') | (df_sing.Level == 'S4')]
    df_sing_AG = df_sing[(df_sing.Level == 'G1') | (df_sing.Level == 'G2')]
    df_sing_FG = df_sing[(df_sing.Level == 'G3') | (df_sing.Level == 'G4') | (df_sing.Level == 'GB')]

    df_coup_AB = df_coup[(df_coup.Level == 'NC') | (df_coup.Level == 'B1') | (df_coup.Level == 'B2')]
    df_coup_FB = df_coup[(df_coup.Level == 'B3') | (df_coup.Level == 'B4')]
    df_coup_AS = df_coup[(df_coup.Level == 'S1') | (df_coup.Level == 'S2')]
    df_coup_FS = df_coup[(df_coup.Level == 'S3') | (df_coup.Level == 'S4')]
    df_coup_AG = df_coup[(df_coup.Level == 'G1') | (df_coup.Level == 'G2')]
    df_coup_FG = df_coup[(df_coup.Level == 'G3') | (df_coup.Level == 'G4') | (df_coup.Level == 'GB')]

    df_contestants = {'Single': [df_sing_AB, df_sing_FB, df_sing_AS, df_sing_FS, df_sing_AG, df_sing_FG],
                      'Couple': [df_coup_AB, df_coup_FB, df_coup_AS, df_coup_FS, df_coup_AG, df_coup_FG]
                     }

    return df_contestants
''' If I go back to saving singles and couples dances in list format each index is a different level

                    while init.dance_dfs_dict[iterator].empty:
                        init.dance_dfs_dict.remove(0)
                        df_list = dance_couples_list
                        query_df = df_list[iterator] # new df to be selecting from 
                    while dance_couples_list[couples_iterator].empty:
                        dance_couples_list.remove(0)
                        df_list = init.dance_dfs_dict
                        query_df = df_list[iterator] # new df to be selecting from
                        level_bp.append(dance_heat_count)
'''


def buildEvent(heats, eventName):
    createParticipantSheets()
    rowindex = 1
    # file = os.getcwd() + eventName + '.xlsx'
    print()
    print("Printing Heats")
    for each in heats:  # For each genre
        # wb.remove_sheet(wb.get_sheet_names()[0])
        for every in heats[each]:  # For every syllabus category
            events = list(heats[each][every].keys())  # Create list of keys
            if events == []:
                continue
            # Create a new directory
            filepath = os.getcwd().replace('\src', "") + "/Output" + "/" + str(each) + "/" + str(every) + "/"
            # filepath = filepath.replace('\\', "/")
            try:
                os.makedirs(filepath)
            except:
                print("Filepath", filepath, "already exists")
            if heats[each][every] == {}:
                continue
            wb = openpyxl.Workbook()
            excelfile = eventName + '_' + str(each) + '_' + str(every) + '.xlsx'
            wb.save(filepath + excelfile)
            rowindex = 2
            for ev in events:  # Loop all events for this genre and syllabus
                print(each, every, ev)
                wb.create_sheet(title=ev)
                wb.active = wb[ev]
                wb.active.page_setup.fitToWidth = 1
                wb.save(filepath+excelfile)
                heatslist = heats[each][every][ev]  # Print all heats for the event
                rosters = heatslist.getRostersList()
                couples_per_floor = heatslist.getCouplesPerFloor()
                count = heatslist.getHeatCount()
                with ExcelWriter(filepath + excelfile, mode='a', if_sheet_exists="overlay") as writer:
                    for heat in rosters:  # Loop over all heats in the Heatlist obj
                        for roomid, room in enumerate(heat.getRoster()):  # For each ballroom print out a list, TODO: may need a -1
                            rowindex += 1
                            startingrow = rowindex
                            # if room == []:
                            #     rowindex += 2
                            # print out each contestant, formatting df to relevant columns
                            for i, contestant in enumerate(room):
                                appendParticipantSheet(heat.getDiv()[roomid], every, ev, roomid, contestant, heat, heatslist)
                                if i == 0 and roomid == 0:
                                    contestant.to_excel(writer, sheet_name=ev, startrow=rowindex-1, columns=init.df_cols, index=False)
                                    rowindex += 1
                                else:
                                    contestant.to_excel(writer, sheet_name=ev, startrow=rowindex-1, columns=init.df_cols, index=False, header=False)
                                rowindex += 1
                            # add in blank rows if data < couples-per-floor
                            if len(room) < couples_per_floor:
                                for i in range(couples_per_floor - len(room)):
                                    rowindex += 1
                        rowindex += 1
                # go over the printed rows and highlight cells for easy identify
                rowindex -= 2
                wb = load_workbook(filename=filepath + excelfile)
                sheet = wb.get_sheet_by_name(ev)
                prev_floor = 1
                # print("Heats", len(heatslist.getRostersList()))
                for i in range(rowindex-1):
                    for col, aline in zip(init.excelcols, init.excelalignments):
                        sheet[col + str(i+1)].alignment = Alignment(horizontal=aline)
                rooms = heatslist.getFloors()
                roomid = 0
                roomindex = 0
                roommax = couples_per_floor + 3
                heatiter = 0
                for i in range(rowindex):
                    # Iterator data
                    index = i + 1
                    if roomindex < roommax:
                        roomindex += 1
                    else:
                        roomindex = 1
                        if roomid < (rooms - 1):
                            roomid += 1
                        else:
                            roomid = 0
                            heatiter = heatiter + 1
                            # wb.save(filepath + excelfile)
                    # If a non contestant data column, add identifier rows, text and color
                    if roomindex == 1 or (roomid == 0 and roomindex == 2):
                        if roomid == 0:
                            roommax = couples_per_floor + 3
                            if roomindex == 1:
                                try:
                                    sheet[init.excelcols[0] + str(index)] = heatslist.getRostersList()[heatiter].getKey()
                                    sheet[init.excelcols[0] + str(index)].fill = PatternFill("solid", start_color="e6bee2")
                                except:
                                    print(index, heatiter, roomid, rowindex)
                            elif roomindex == 2:
                                # Replace lvl # with it's corresponding name EX: AB, FB AS-FS
                                try:
                                    div = heatslist.getRostersList()[heatiter].getDiv()[roomid]
                                except:
                                    print(index, heatiter, roomid, rowindex)
                                if div == []:
                                    sheet[init.excelcols[0] + str(index)] = 'Floor ' + str(roomid + 1) + " Not used"  # ILLEGAL_CHARACTERS_RE.sub(r'', str(room))
                                    sheet[init.excelcols[0] + str(index)].fill = PatternFill("solid", start_color="a9ebba")
                                    sheet.merge_cells(init.excelcols[0] + str(index) + ":" + init.excelcols[-1] + str(index))
                                    sheet[init.excelcols[0] + str(index)].alignment = Alignment(horizontal='left')
                                    continue
                                if div[0] == "S":
                                    for lev in init.lvl_conversion:
                                        if lev in div:
                                            div[div.index(lev)] = heatslist.getEventLvlSingles()[lev]
                                    for age in heatslist.getEventAgesSingles():
                                        if age in div:
                                            printagelist = heatslist.getEventAgesSingles().copy()
                                            printagelist.insert(0, 17)
                                            age_index = printagelist.index(div[div.index(age)])
                                            if age_index == printagelist[-1]:  # If the last bracket make it '86+'
                                                div[div.index(age)] = str(printagelist[age_index - 1] + 1) + "+"
                                            else:
                                                div[div.index(age)] = str(printagelist[age_index - 1] + 1) + "-" + str(printagelist[age_index])
                                if div[0] == "C":
                                    for lev in init.lvl_conversion:
                                        if lev in div:
                                            div[div.index(lev)] = heatslist.getEventLvlCouples()[lev]
                                    for age in heatslist.getEventAgesCouples():
                                        if age in div:
                                            printagelist = heatslist.getEventAgesCouples().copy()
                                            printagelist.insert(0, 17)
                                            age_index = printagelist.index(div[div.index(age)])
                                            if age_index == printagelist[-1]:  # If the last bracket make it '86+'
                                                div[div.index(age)] = str(printagelist[age_index - 1] + 1) + "+"
                                            else:
                                                div[div.index(age)] = str(printagelist[age_index - 1] + 1) + "-" + str(printagelist[age_index])
                                printstr = ""
                                for d in div:
                                    printstr = printstr + str(d) + ", "
                                sheet[init.excelcols[0] + str(index)] = 'Floor ' + str(roomid + 1) + " " + printstr[:-2]  # ILLEGAL_CHARACTERS_RE.sub(r'', str(room))
                                sheet[init.excelcols[0] + str(index)].fill = PatternFill("solid", start_color="a9ebba")
                        elif roomid > 0 and roomindex == 1:
                            roommax = couples_per_floor + 1
                            # Replace lvl # with it's corresponding name EX: AB, FB AS-FS
                            div = heatslist.getRostersList()[heatiter].getDiv()[roomid]
                            if div == []:
                                sheet[init.excelcols[0] + str(index)] = 'Floor ' + str(roomid + 1) + " Not used"  # ILLEGAL_CHARACTERS_RE.sub(r'', str(room))
                                sheet[init.excelcols[0] + str(index)].fill = PatternFill("solid", start_color="a9ebba")
                                sheet.merge_cells(init.excelcols[0] + str(index) + ":" + init.excelcols[-1] + str(index))
                                sheet[init.excelcols[0] + str(index)].alignment = Alignment(horizontal='left')
                                continue
                            if div[0] == "S":
                                for lev in init.lvl_conversion:
                                    if lev in div:
                                       div[div.index(lev)] = heatslist.getEventLvlSingles()[lev]
                                for age in heatslist.getEventAgesSingles():
                                    if age in div:
                                        printagelist = heatslist.getEventAgesSingles().copy()
                                        printagelist.insert(0, 17)
                                        age_index = printagelist.index(div[div.index(age)])
                                        if age_index == printagelist[-1]:  # If the last bracket make it '86+'
                                            div[div.index(age)] = str(printagelist[age_index - 1] + 1) + "+"
                                        else:
                                            div[div.index(age)] = str(printagelist[age_index - 1] + 1) + "-" + str(printagelist[age_index])
                            if div[0] == "C":
                                for lev in init.lvl_conversion:
                                    if lev in div:
                                        div[div.index(lev)] = heatslist.getEventLvlCouples()[lev]
                                for age in heatslist.getEventAgesCouples():
                                    if age in div:
                                        printagelist = heatslist.getEventAgesCouples().copy()
                                        printagelist.insert(0, 17)
                                        age_index = printagelist.index(div[div.index(age)])
                                        if age_index == printagelist[-1]: # If the last bracket make it '86+'
                                            div[div.index(age)] = str(printagelist[age_index-1] + 1) + "+"
                                        else:
                                            div[div.index(age)] = str(printagelist[age_index-1] + 1) + "-" + str(printagelist[age_index])
                            printstr = ""
                            for d in div:
                                printstr = printstr + str(d) + ", "
                            sheet[init.excelcols[0] + str(index)] = 'Floor ' + str(roomid + 1) + " " + printstr[:-2] # ILLEGAL_CHARACTERS_RE.sub(r'', str(room))
                            sheet[init.excelcols[0] + str(index)].fill = PatternFill("solid", start_color="a9ebba")
                        sheet.merge_cells(init.excelcols[0]+str(index)+":"+init.excelcols[-1]+str(index))
                        sheet[init.excelcols[0] + str(index)].alignment = Alignment(horizontal='left')
                        # prev_floor = index
                        # wb.save(filepath+excelfile)
                        continue
                    # If Contestant Data header row
                    if roomid == 0 and roomindex == 3:
                        for col in init.excelcols:
                            sheet[col + str(index)].alignment = Alignment(horizontal='center')

                    # Color code the lvl metals
                    if sheet["A" + str(index)].value == None:
                        sheet[idcol + str(index)].fill = PatternFill("solid", start_color="f50031")
                    idcol = init.excelcols[0]
                    if sheet[idcol+str(index)].value == "C":
                        sheet[idcol+str(index)].fill = PatternFill("solid", start_color="a2c2f5")
                    elif sheet[idcol+str(index)].value == "F":
                        sheet[idcol+str(index)].fill = PatternFill("solid", start_color="f2a7aa")
                    # level color code
                    levcol = init.excelcols[7]
                    if sheet[levcol + str(index)].value is not None:
                        # Bronze/NC
                        if sheet[levcol+str(index)].value[0] == "B" or sheet[levcol+str(index)].value[0] == "N":
                            if sheet[levcol+str(index)].value[1] == "1" or sheet[levcol+str(index)].value[1] == "2" or sheet[levcol+str(index)].value[1] == "C":
                                sheet[levcol+str(index)].fill = PatternFill("solid", start_color="eddcca")
                            if sheet[levcol+str(index)].value[1] == "3" or sheet[levcol+str(index)].value[1] == "4":
                                sheet[levcol+str(index)].fill = PatternFill("solid", start_color="f2ba7e")
                        # Silver
                        if sheet[levcol+str(index)].value[0] == "S":
                            if sheet[levcol+str(index)].value[1] == "1" or sheet[levcol+str(index)].value[1] == "2":
                                sheet[levcol+str(index)].fill = PatternFill("solid", start_color="e0dfde")
                            if sheet[levcol+str(index)].value[1] == "3" or sheet[levcol+str(index)].value[1] == "4":
                                sheet[levcol+str(index)].fill = PatternFill("solid", start_color="a39e99")
                        # Gold
                        if sheet[levcol+str(index)].value[0] == "G":
                            if sheet[levcol+str(index)].value[1] == "1" or sheet[levcol+str(index)].value[1] == "2":
                                sheet[levcol+str(index)].fill = PatternFill("solid", start_color="f0eec5")
                            if sheet[levcol+str(index)].value[1] == "3" or sheet[levcol+str(index)].value[1] == "4" or sheet[levcol+str(index)].value[1] == "B":
                                sheet[levcol+str(index)].fill = PatternFill("solid", start_color="f5ee71")
                # Set column dimensions
                for col, dim in zip(init.excelcols, init.exceldimensions):
                    sheet.column_dimensions[col].width = dim
            wb.save(filepath+excelfile)
    print()
    print("Creating Personal Heat Sheets")
    # Loop over all participant sheets and print them
    participants = list(init.participantsheets.keys())
    for each in participants:
        df = init.participantsheets[each]
        # Create a new directory
        filepath = os.getcwd().replace('\src', "") + "/Output" + "/" + "Personal Sheets" + "/" + str(each) + "/"
        # filepath = filepath.replace('\\', "/")
        try:
            os.makedirs(filepath)
        except:
            pass
            # print("Filepath", filepath, "already exists")
        wb = openpyxl.Workbook()
        excelfile = str(each) + '.xlsx'
        sheet = wb.active
        sheet.merge_cells(init.participantsheet_excelcols[0] + str(1) + ":" + init.participantsheet_excelcols[-1] + str(1))
        sheet[init.participantsheet_excelcols[0] + str(1)].value = str(each) + "  " + init.df_Dnum["First Name"].loc[init.df_Dnum["Dancer #"] == each].values[0] + ", " + init.df_Dnum["Last Name"].loc[init.df_Dnum["Dancer #"] == each].values[0]
        # Set column dimensions
        for col, dim in zip(init.participantsheet_excelcols, init.participantsheet_exceldimensions):
            sheet.column_dimensions[col].width = dim
        for i in range(df.shape[0]+1):
            index = i + 2
            for col, aline in zip(init.participantsheet_excelcols, init.participantsheet_excelalignments):
                sheet[col + str(index)].alignment = Alignment(horizontal=aline)
        wb.save(filepath + excelfile)
        with ExcelWriter(filepath + excelfile, mode='a', if_sheet_exists="overlay") as writer:
            df.to_excel(writer, sheet_name='Sheet', startrow=1, index=False)



def getContestantConflictList(dance_df, inst, contestants):
    """Finds the all contestants that will conflict for the current heat roster given the instructor inst

       Parameters
       ----------
       dance_df : Pandas DataFrame
           Current pool of single contestants
       inst : int
           Instructor Randomly picked from list of instructors possible for the level
       contestants : list[[int]]
           List of 6 lists holding all contestants that are placed in current heat
           List of 6 lists holding all contestants that are placed in current heat

       Returns
       -------
       list[int]
           a list of Dancer #'s that will conflict with the current heat roster
       """
    dupe_con = []
    for row, singles in dance_df.iterrows():
        # Set column variables based on single type
        if singles["type id"] == "L":
            contestant_col = "Lead Dancer #"
            inst_col = "Follow Dancer #"
            inst_fname = "Follow First Name"
            inst_lname = "Follow Last Name"
        elif singles["type id"] == "F":
            contestant_col = "Follow Dancer #"
            inst_col = "Lead Dancer #"
            inst_fname = "Lead First Name"
            inst_lname = "Lead Last Name"
        else:
            raise Exception("Type id for " + singles + " is invalid")
        # Retrofit the instructor list data
        if type(singles["Instructor Dancer #'s"]) == int:
            tmp = [singles["Instructor Dancer #'s"]]
        else:
            tmp = [int(x) for x in singles["Instructor Dancer #'s"].split(";")]
        singles["Instructor Dancer #'s"] = tmp
        del tmp
        for num in singles["Instructor Dancer #'s"]:
            if num == inst:
                for each in contestants:
                    if each.count(singles[contestant_col]) > 0:
                        dupe_con.append(each)
    return dupe_con


def getContestantFreeList(dance_df, inst, contestants):
    """Finds the all contestants that are free for the current heat roster given the instructor inst

       Parameters
       ----------
       dance_df : Pandas DataFrame
           Current pool of single contestants
       inst : int
           Instructor Randomly picked from list of instructors possible for the level
       contestants : list[[int]]
           List of 6 lists holding all contestants that are placed in current heat
           List of 6 lists holding all contestants that are placed in current heat

       Returns
       -------
       list[int]
           a list of Dancer #'s that are free with the current heat roster
       """
    free_con = []
    for row, singles in dance_df.iterrows():
        # Set column variables based on single type
        if singles["type id"] == "L":
            contestant_col = "Lead Dancer #"
            inst_col = "Follow Dancer #"
            inst_fname = "Follow First Name"
            inst_lname = "Follow Last Name"
        elif singles["type id"] == "F":
            contestant_col = "Follow Dancer #"
            inst_col = "Lead Dancer #"
            inst_fname = "Lead First Name"
            inst_lname = "Lead Last Name"
        else:
            raise Exception("Type id for " + singles + " is invalid")
        for num in singles["Instructor Dancer #'s"]:
            if num == inst:
                for each in contestants:
                    if each.count(singles[contestant_col]) == 0:
                        free_con.append(each)
    return free_con

def deleteEmpty(dance_dfs):
    keylist = list(dance_dfs.keys())
    for each in keylist:
        if type(dance_dfs[each]) is dict:
            deleteEmpty(dance_dfs[each])
            if len(dance_dfs[each]) == 0:  # If dict is now empty, delete it
                del dance_dfs[each]
        elif dance_dfs[each].empty:
            del dance_dfs[each]
    return dance_dfs


def buildInst2SingTree(dance_dfs, inst2sing_tree, ev):
    """Builds the dictionary tree of all instructors and thier count in that division

           Parameters
           ----------
           dance_df : Pandas DataFrame
               Current pool of single contestants
           inst_tree : Python Dictionary
               Tree of dictionaries key'd by the division metrics

           Returns
           -------
           dict
              tree of all instructor counts key'd by thier division
           """
    keylist = list(dance_dfs.keys())
    for each in keylist:
        if type(dance_dfs[each]) is dict:
            inst2sing_tree[each] = {}
            buildInst2SingTree(dance_dfs[each], inst2sing_tree[each], ev)
        else:
            singles_only = dance_dfs[each][(dance_dfs[each]['type id'] == 'L') | (dance_dfs[each]['type id'] == 'F')]
            if not singles_only.empty:
                inst2sing_tree[each] = {}
                # Find unique instructors for this division
                for row, data in singles_only.iterrows():  # Iterate down all rows of Single df
                    if data["type id"] == "L":
                        contestant_col = "Lead Dancer #"
                        inst_col = "Follow Dancer #"
                        inst_fname = "Follow First Name"
                        inst_lname = "Follow Last Name"
                    elif data["type id"] == "F":
                        contestant_col = "Follow Dancer #"
                        inst_col = "Lead Dancer #"
                        inst_fname = "Lead First Name"
                        inst_lname = "Lead Last Name"
                    if type(data["Instructor Dancer #'s"]) == int:
                        tmp = [data["Instructor Dancer #'s"]]
                    else:
                        tmp = [int(x) for x in data["Instructor Dancer #'s"].split(";")]
                    data["Instructor Dancer #'s"] = tmp
                    del tmp
                    for num in data["Instructor Dancer #'s"]:  # Iterate through all #'s in instructor lists
                        # singles for that instructor tree
                        if inst2sing_tree[each].get(num) is None:
                            inst2sing_tree[each][num] = [data[contestant_col]]
                        if data[contestant_col] not in inst2sing_tree[each][num]:
                            inst2sing_tree[each][num].append(data[contestant_col])

    return inst2sing_tree


def createParticipantSheets():
    # Loop over each data set and create an empty df where there specific heat data will be added
    # only add one if not already there
    for row, data in init.df_sing.iterrows():
        dancer = data.loc["Dancer #"]
        if init.participantsheets.get(dancer) is None:
            init.participantsheets[dancer] = pd.DataFrame(init.participantsheetcols)

    for row, data in init.df_inst.iterrows():
        dancer = data.loc["Dancer #"]
        if init.participantsheets.get(dancer) is None:
            init.participantsheets[dancer] = pd.DataFrame(init.participantsheetcols)

    for row, data in init.df_coup.iterrows():
        dancer = data.loc["Lead Dancer #"]
        if init.participantsheets.get(dancer) is None:
            init.participantsheets[dancer] = pd.DataFrame(init.participantsheetcols)
        dancer = data.loc["Follow Dancer #"]
        if init.participantsheets.get(dancer) is None:
            init.participantsheets[dancer] = pd.DataFrame(init.participantsheetcols)


def appendParticipantSheet(div, syllabus, ev, roomid, roster_entry, heat, heatslist):
    typeid = roster_entry.loc[0, "type id"]
    if typeid == "C":
        leads_df = init.df_coup
        lead_col = "Lead Dancer #"
        follows_df = init.df_coup
        follow_col = "Follow Dancer #"
    elif typeid == "L":
        leads_df = init.df_sing
        lead_col = "Dancer #"
        follows_df = init.df_inst
        follow_col = "Dancer #"
    elif typeid == "F":
        leads_df = init.df_inst
        lead_col = "Dancer #"
        follows_df = init.df_sing
        follow_col = "Dancer #"

    if div[0] == "S":
        eventlvls = heatslist.getEventLvlSingles()
        eventages = heatslist.getEventAgesSingles()
    elif div[0] == "C":
        eventlvls = heatslist.getEventLvlCouples()
        eventages = heatslist.getEventAgesCouples()

    for lev in init.lvl_conversion:
        if lev in div:
            div[div.index(lev)] = eventlvls[lev]
    for age in eventages:
        if age in div:
            printagelist = eventages.copy()
            printagelist.insert(0, 17)
            age_index = printagelist.index(div[div.index(age)])
            if age_index == printagelist[-1]:  # If the last bracket make it '86+'
                div[div.index(age)] = str(printagelist[age_index - 1] + 1) + "+"
            else:
                div[div.index(age)] = str(printagelist[age_index - 1] + 1) + "-" + str(printagelist[age_index])
    printstr = ""
    for d in div:
        printstr = printstr + str(d) + ", "
    # Set Participant df data
    dancer = roster_entry.loc[0, lead_col]
    partner = roster_entry.loc[0, follow_col]
    partner_n = roster_entry.loc[0, "Follow First Name"] + " " + roster_entry.loc[0, "Follow Last Name"]
    participantsheetentry = {"Day": [1], "Heat #": [heat.getKey()], "Floor": [roomid], "Partner #": [partner], "Partner Name": [partner_n], "Event": [ev], "Syllabus": [syllabus],
                            "Division": [printstr[:-2]]}
    participantsheetentry = pd.DataFrame(participantsheetentry)
    # participantsheetentry = participantsheetentry.astype({"Day": int, "Heat #": int, "Floor": int, "Partner #": int})
    participantsheetentry = participantsheetentry.astype({"Day": int, "Floor": int, "Partner #": int})
    if not init.participantsheets[dancer].empty:
        pass
    init.participantsheets[dancer] = pd.concat([init.participantsheets[dancer], participantsheetentry])
    # init.participantsheets[dancer] = init.participantsheets[dancer].astype({"Day": int, "Heat #": int, "Floor": int, "Partner #": int})
    init.participantsheets[dancer] = init.participantsheets[dancer].astype({"Day": int, "Floor": int, "Partner #": int})

    # Set Participant df data
    dancer = roster_entry.loc[0, follow_col]
    partner = roster_entry.loc[0, lead_col]
    partner_n = roster_entry.loc[0, "Lead First Name"] + " " + roster_entry.loc[0, "Lead Last Name"]
    participantsheetentry = {"Day": [1], "Heat #": [heat.getKey()], "Floor": [roomid], "Partner #": [partner], "Partner Name": [partner_n], "Event": [ev],
                             "Syllabus": [syllabus],
                             "Division": [printstr[:-2]]}
    participantsheetentry = pd.DataFrame(participantsheetentry)
    # participantsheetentry = participantsheetentry.astype({"Day": int, "Heat #": int, "Floor": int, "Partner #": int})
    participantsheetentry = participantsheetentry.astype({"Day": int, "Floor": int, "Partner #": int})
    if not init.participantsheets[dancer].empty:
        pass
    init.participantsheets[dancer] = pd.concat([init.participantsheets[dancer], participantsheetentry])
    # init.participantsheets[dancer] = init.participantsheets[dancer].astype({"Day": int, "Heat #": int, "Floor": int, "Partner #": int})
    init.participantsheets[dancer] = init.participantsheets[dancer].astype({"Day": int, "Floor": int, "Partner #": int})


def pickDfs(ev, dance_dfs, inst_tree, floors, div, eventages_s, eventages_c, couples_per_floor):
    print()
    l_keys = [0, 1, 2, 3, 4, 5]

    if div.count('T') == 0 and div.count('t') == 0:
        start = "A"
    # elif "S" in dance_dfs:
    #     start = "S"
    else:
        start = "C"

    shell = dance_dfs[start]
    picked_keys = []
    picked_keys.append(start)
    # TODO how to handel lead and follow 3 or 4 split with singles
    while type(shell) is dict:
        # pick = random.choice(list(shell.keys()))  # random choice
        pick = list(shell.keys())[0]
        picked_keys.append(pick)
        shell = shell[pick]

    # If only one floor return the key
    if len([picked_keys]) == floors:
        return [picked_keys]
    unideal = []  # keys that do not meet initial conditions but if no other options will choose one of these
    # If picked df is a single
    if picked_keys[0] == "S":
        iter = 1  # Iterator to track key index
        # if divided by L and F, set labels
        if ("S" in div) or ("s" in div):
            same_s = "same_s"
            opp_s = "opp_s"
            iter += 1
        else:  # assignment means nothing for else case only to stop compile error, since this key DNE
            same_s = "Follow"
            opp_s = "Lead"

        # If divided by level set the same level
        if ("L" in div) or ("l" in div):
            same_l = picked_keys[iter]
            diff_l = "diff_l"
            lev_index = iter
            iter += 1
        else:  # assignment means nothing for else case only to stop compile error, since this key DNE
            same_l = 0
            diff_l = 0

        # If divided by age set the same age bracket
        if ("A" in div) or ("a" in div):
            same_a = picked_keys[iter]
            diff_a = "diff_a"
            age_index = iter
            iter += 1
        else:  # assignment means nothing for else case only to stop compile error, since this key DNE
            same_a = 0
            diff_a = 0

        #        Opposite Gender
        prio_s = [["S", opp_s, same_l, same_a], ["S", opp_s, same_l, diff_a], ["S", opp_s, diff_l, same_a],
                  ["S", opp_s, diff_l, diff_a],
        #         Same Gender
                  ["S", same_s, same_l, diff_a], ["S", same_s, diff_l, same_a], ["S", same_s, diff_l, diff_a],
        #         Couples
                  ["C", same_l, same_a], ["C", same_l, diff_a], ["C", diff_l, same_a], ["C", diff_l, diff_a]]
        all_poss = []  # holds all possible keys to be selected, used in the case of different level or age
        picked_keys = [picked_keys]
        prio = 0
        while prio < len(prio_s):
            if dance_dfs.get(prio_s[prio][0]) is None:
                prio += 1
                continue
            poss_key = []
            iter = 1
            # Take out unused key slots
            if div.count('T') != 0 or div.count('t') != 0:
                poss_key.append(prio_s[prio][0])
            else:
                poss_key.append("A")
            if (div.count('s') != 0 or div.count('S') != 0) and poss_key[0] == "S":
                if prio_s[prio][1] == "opp_s":
                    if picked_keys[-1][iter] == "Lead":
                        poss_key.append("Follow")
                    else:
                        poss_key.append("Lead")
                else:
                    if picked_keys[-1][iter] == "Follow":
                        poss_key.append("Follow")
                    else:
                        poss_key.append("Lead")
                iter += 1
            if div.count('L') != 0 or div.count('l') != 0:
                if prio_s[prio][2] != "diff_l":
                    poss_key.append(prio_s[prio][2])
                    all_poss.append(poss_key)
                else:  # If different level make a key with all possible levels
                    for lev in l_keys:  # For all level keys
                        if lev != picked_keys[0][lev_index]:  # if not the same level, create a new possible key
                            poss_key.append(lev)
                            all_poss.append(poss_key)
                            poss_key = poss_key[:-1]
                iter += 1
            if div.count('A') != 0 or div.count('a') != 0:
                if len(all_poss) == 0:  # in case 'l' split is not part of this Event
                    all_poss.append(poss_key)
                if prio_s[prio][3] != "diff_a":
                    for key in all_poss:
                        key.append(prio_s[prio][3])
                else:  # If different age make a key with all possible ages
                    curr_all_poss = len(all_poss)
                    if prio_s[prio][0] == "S":  # Make sure to give the correct age brackets based on the current possible key(s)
                        ages_to_use = eventages_s
                    else:
                        ages_to_use = eventages_c
                    agelist = ages_to_use.copy()
                    agelist.remove(picked_keys[0][age_index])
                    # iter = 0
                    for age in agelist:
                        for j in range(curr_all_poss):
                            poss_key = all_poss[j][:]
                            poss_key.append(age)
                            all_poss.append(poss_key)
                    # Remove the all_poss with no age
                    all_poss = all_poss[curr_all_poss:]
            # Use all_poss to find out if that combo exists in the data tree
            picked = False
            for every in all_poss:
                for i, key in enumerate(every):
                    if i == 0:
                        tmp = dance_dfs[key]
                        continue
                    if key not in tmp:
                        break
                    if type(tmp[key]) is dict:
                        tmp = tmp[key]
                    elif every not in picked_keys:  # add to picked list only if that key combo is not in already
                        # if added is a single
                        if every[0] == "S":
                            cop = []
                            for key in picked_keys:
                                cop.append(key[:])
                            # Sanity check the levels in question
                            unique = findUnique(inst_tree, cop, every)
                            count = findInstCount(inst_tree, every)
                            # find # of 'singles' rooms
                            sfloors = 0
                            for room in picked_keys:
                                if room[0] == "S":
                                    sfloors += 1
                            if every[0] == "S":
                                sfloors += 1
                            # if (unique >= (len(picked_keys) * couples_per_floor)-3) and (count > (couples_per_floor-2)):
                            if unique >= (sfloors * couples_per_floor)-2:
                                picked_keys.append(every)
                                picked = True
                                break
                            else:
                                unideal.append([every, count])
                        else:  # if not a single
                            count = findContestantCount(dance_dfs, every, ev)
                            if count[0] >= couples_per_floor-2:
                                picked_keys.append(every)
                                picked = True
                            else:
                                unideal.append([every, count[1]])
                if picked:  # If picked break
                    picked = False
                    prio = 1  # So I don't skip other gender prio options
                    break
                if every == all_poss[-1]:  # If at the end of options for this priority bracket move to another
                    prio += 1
            all_poss.clear()  # Once all possible keys in priority slot are not picked, clear the list
            if len(picked_keys) == floors:
                break
                
    # If picked df is a couple
    elif picked_keys[0] == "C" or picked_keys[0] == "A":
        iter = 1  # Iterator to track key index
        # If divided by level set the same level
        if ("L" in div) or ("l" in div):
            same_l = picked_keys[iter]
            diff_l = "diff_l"
            lev_index = iter
            iter += 1
        else:  # assignment means nothing for else case only to stop compile error, since this key DNE
            same_l = 0
            diff_l = 0

        # If divided by age set the same age bracket
        if ("A" in div) or ("a" in div):
            same_a = picked_keys[iter]
            diff_a = "diff_a"
            age_index = iter
            iter += 1
        else:  # assignment means nothing for else case only to stop compile error, since this key DNE
            same_a = 0
            diff_a = 0

        # Pick priority list depending on starting key
        if picked_keys[0] == "C":
            #         Couples
            prio_c = [["C", same_l, diff_a], ["C", diff_l, same_a], ["C", diff_l, diff_a],
            #         Singles
            #           ["S", "single", same_l, same_a], ["S", "single", same_l, diff_a], ["S", "single", diff_l, same_a],
            #           ["S", "single", diff_l, diff_a]
                      ]
        elif picked_keys[0] == "A":
            prio_c = [["A", same_l, diff_a], ["A", diff_l, same_a], ["A", diff_l, diff_a]]

        all_poss = []  # holds all possible keys to be selected, used in the case of different level or age
        picked_keys = [picked_keys]
        prio = 0
        while prio < len(prio_c):

            poss_key = []
            iter = 1
            if dance_dfs.get(prio_c[prio][0]) is None:
                prio += 1
                continue
            # Take out unused key slots
            if div.count('T') != 0 or div.count('t') != 0:
                poss_key.append(prio_c[prio][0])
            else:
                poss_key.append("A")
            if (div.count('s') != 0 or div.count('S') != 0) and poss_key[0] == "S":
                poss_key.append("Lead")
                all_poss.append(poss_key)
                poss_key = poss_key[:-1]
                poss_key.append("Follow")
                all_poss.append(poss_key)
                iter += 1
            if div.count('L') != 0 or div.count('l') != 0:
                if len(all_poss) == 0:  # in case 's' split is not part of this event
                    all_poss.append(poss_key)
                if prio_c[prio][1] != "diff_l":
                    for key in all_poss:
                        key.append(prio_c[prio][1])
                else:  # If different level make a key with all possible levels
                    curr_all_poss = len(all_poss)
                    for lev in l_keys:  # duplicate current all_poss per different level
                        if lev != picked_keys[0][lev_index]:
                            for j in range(curr_all_poss):
                                poss_key = all_poss[j][:]
                                poss_key.append(lev)
                                all_poss.append(poss_key)
                    # remove first 2 keys that have no age
                    all_poss = all_poss[curr_all_poss:]
                iter += 1
            if div.count('A') != 0 or div.count('a') != 0:
                if len(all_poss) == 0:  # In case 's' split is not part of this
                    all_poss.append(poss_key)
                if prio_c[prio][2] != "diff_a":
                    for key in all_poss:
                        key.append(prio_c[prio][2])
                else:  # If different age make a key with all existing key with all ages
                    curr_all_poss = len(all_poss)
                    if prio_c[prio][0] == "S":  # Make sure to give the correct age brackets for selection, based on the possible keys
                        ages_to_use = eventages_s
                    else:
                        ages_to_use = eventages_c
                    agelist = ages_to_use.copy()
                    agelist.remove(picked_keys[0][age_index])
                    for age in agelist:
                        for j in range(curr_all_poss):
                            poss_key = all_poss[j][:]
                            poss_key.append(age)
                            all_poss.append(poss_key)
                    # Remove the all_poss with no age
                    all_poss = all_poss[curr_all_poss:]

            # Use all_poss to find out if that combo exists in the data tree
            picked = False
            print("Checking all possible keys", all_poss)
            for every in all_poss:
                for i, key in enumerate(every):
                    if i == 0:
                        tmp = dance_dfs[key]
                        continue
                    if key not in tmp:
                        break
                    if type(tmp[key]) is dict:
                        tmp = tmp[key]
                    elif every not in picked_keys:  # add to picked list only if that key combo is not in already
                        # if added is a single
                        if every[0] == "S":
                            cop = []
                            for key in picked_keys:
                                cop.append(key[:])
                            # Sanity check the levels in question
                            unique = findUnique(inst_tree, cop, every)
                            count = findInstCount(inst_tree, every)
                            # find # of 'singles' rooms
                            sfloors = 0
                            for room in picked_keys:
                                if room[0] == "S":
                                    sfloors += 1
                            if every[0] == "S":
                                sfloors += 1
                            # if (unique >= (len(picked_keys) * couples_per_floor)-3) and (count > (couples_per_floor-2)):
                            if unique >= (sfloors * couples_per_floor) - 2:
                                picked_keys.append(every)
                                picked = True
                                break
                            else:
                                unideal.append([every, count])
                        else:  # if not a single
                            print("Checking if", every, "can be added")
                            count = findContestantCount(dance_dfs, every, ev)
                            if count[0] >= (couples_per_floor - 2):
                                picked_keys.append(every)
                                picked = True
                            else:
                                print("added an unideal heat", every)
                                unideal.append([every, count[1]])
                if picked:  # If picked break
                    picked = False
                    prio = 0  # So I don't skip other gender prio options
                    break
                if every == all_poss[-1]:  # If at the end of options for this priority bracket move to another
                    prio += 1
            all_poss.clear()  # Once all possible keys in priority slot are not picked, clear the list
            if len(picked_keys) == floors:
                break

        # If all possible keys are tried and not filled the floors try the unideal heats
        while len(picked_keys) != floors:
            if len(unideal) == 0:
                break
            largest = 0
            largest_index = 0
            for i, each in enumerate(unideal):
                if each[1] > largest:
                    largest = each[1]
                    largest_index = i
            print("Using Unideal heat", unideal[largest_index][0])
            picked_keys.append(unideal[largest_index][0])
            del unideal[largest_index]
    if len(picked_keys) < floors:
        pass
    return picked_keys


def findUnique(inst_tree, picked_keys, newkey):

    instructors = []

    picked_keys.append(newkey)
    for each in picked_keys:
        if each[0] != "C":
            tmp = inst_tree[each[0]]
            i = 1
            while type(tmp) is dict and i < len(each):
                tmp = tmp[each[i]]
                i += 1
            instructors = list(tmp.keys()) + instructors

    return len(set(instructors))

def findInstCount(inst_tree, newkey):
    for i, each in enumerate(newkey):
        if i == 0:
            tmp = inst_tree[each]
            continue
        tmp = tmp[each]

    return len(tmp.keys())

def findContestantCount(dance_dfs, newkey, ev):
    try:
        for i, each in enumerate(newkey):
            if i == 0:
                tmp = dance_dfs[each]
                continue
            tmp = tmp[each]
        return [tmp.shape[0], tmp[ev].sum()]
    except:
        print(newkey, "No longer in pool")

def PoachPrevHeatsSingles(roomid, div, dance_df, heat, heatlist, acceptablecouples):
    counter = 0  # Counter for termination condition
    couples_per_floor = heatlist.getCouplesPerFloor()
    poachlist = []
    poachsingles = []
    # While current heat's selection room is less than half the size of a full room
    for heat_index, each in enumerate(heatlist.getRostersList()):
        # Check if heat has division
        if div not in each.getDiv():
            continue
        # # Find an entry in the previous heat room that has no conflicts with current heat
        # if each.getDiv().count(div) > 1:
        #     past_heat_iter = each.getDiv().count(div)
        # else:
        #     past_heat_iter = 1
        # start_at = 0
        # for j in range(past_heat_iter):  # Catches any heat that has multi same division floors
        swapping_room = each.getDiv().index(div, start_at)
        start_at = swapping_room
        index_iter = 0
        index_2_swap = -1
        for contestant, inst in zip(each.getSingles()[swapping_room], each.getInstructors()[swapping_room]):
            for i, singles in enumerate(heat.getSingles()):
                if contestant in singles:  # make sure the conflict is not because it has some inst trying to be freed
                    dup = True
                    break
            for i, instructors in enumerate(heat.getInstructors()):
                if contestant in instructors:  # make sure the conflict is not because it has some inst trying to be freed
                    raise Exception("Couples Conflict Contestant", contestant, "in Singles Instructor list")
                    # conflict to check the previous heat entry nth conflict
                    # resolverconflict = ResolverConflictItemSingle(3, conflict_div, heat_index,swapping_room, placed_inst[swapping_room].index(inst), instructors_in_heat, singles_in_heat, inst, conflict, [swapper, swappee])
                    # resolverLog.addConflict(resolverconflict, con_num)
                    # print("Swappee conflict in heat,room,index", heat_index, swapping_room, placed_inst[swapping_room].index(inst), 'inst', inst)
                if inst in instructors:
                    dup = True
                    break
            for i, couples in enumerate(heat.getCouples()):
                if contestant in couples:
                    dup = True
                    break
            if not dup:
                index_2_swap = int((index_iter - 1) / 2)
                s = contestant
                if [heat_index, swapping_room, index_2_swap] not in poachlist and (s not in poachsingles):  # Check this couple is not in the list already
                    poachlist.append([heat_index, swapping_room, index_2_swap])
                    poachsingles.append(s)
                    break
            dup = False
            index_iter += 1

    if (len(poachlist) + len(heat.getSingles()[roomid])) < acceptablecouples:
        for i, contestant in enumerate(reversed(heat.getSingles()[roomid])):
            index = len(heat.getSingles()[roomid]) - 1
            tmp = heat.stealEntry(roomid, index)
            dance_df = pd.concat([dance_df, tmp])
        print("Backfilling", div)
        backfill(dance_df, div, heatlist, couples_per_floor, init.ev)
    else:
        print("Poaching from heats", div, poachlist)
        heatlen = len(heat.getRoster()[roomid])
        counter = 0
        while heatlen < acceptablecouples:
            for poach in poachlist:
                if heatlen == acceptablecouples:
                    break
                # poach = random.choice(poachlist)
                heat2poach = heatlist.getRostersList()[poach[0]]
                if len(heat2poach.getRoster()[poach[1]]) >= couples_per_floor and counter == 0:
                    print("Contestant", heat2poach.getCouples()[poach[1]][poach[2]], " stolen from heat, room",
                          poach[0], poach[1])
                    tmp = heat2poach.stealEntry(poach[1], poach[2])
                    heat.addEntry(tmp, roomid)
                    poachlist.remove(poach)
                    heatlen += 1
                elif counter >= 1:
                    print("Contestant", heat2poach.getCouples()[poach[1]][poach[2]], " stolen from heat, room",
                          poach[0], poach[1])
                    tmp = heat2poach.stealEntry(poach[1], poach[2])
                    heat.addEntry(tmp, roomid)
                    poachlist.remove(poach)
                    heatlen += 1
            counter += 1


def PoachPrevHeatsCouples(roomid, div, dance_df, heat, heatlist, acceptablecouples):
    couples_per_floor = heatlist.getCouplesPerFloor()
    poachlist = []
    poachcouples = []
    # While current heat's selection room is less than half the size of a full room
    for heat_index, each in enumerate(heatlist.getRostersList()):
        # Check if heat has division
        if div not in each.getDiv():
            continue
        # Find an entry in the previous heat room that has no conflicts with current heat
        if each.getDiv().count(div) > 1:
            past_heat_iter = each.getDiv().count(div)
        else:
            past_heat_iter = 1
        start_at = 0

        swapping_room = each.getDiv().index(div, start_at)
        start_at = swapping_room
        index_iter = 0
        index_2_swap = -1
        lookin = each.getCouples()[swapping_room]
        dup = False
        for contestant in lookin:
            for i, singles in enumerate(heat.getSingles()):
                if contestant in singles:  # make sure the conflict is not because it has some inst trying to be freed
                    dup = True
                    break
            for i, instructors in enumerate(heat.getInstructors()):
                if contestant in instructors:  # make sure the conflict is not because it has some inst trying to be freed
                    raise Exception("Couples Conflict Contestant", contestant, "in Singles Instructor list")
                    # conflict to check the previous heat entry nth conflict
                    # resolverconflict = ResolverConflictItemSingle(3, conflict_div, heat_index,swapping_room, placed_inst[swapping_room].index(inst), instructors_in_heat, singles_in_heat, inst, conflict, [swapper, swappee])
                    # resolverLog.addConflict(resolverconflict, con_num)
                    # print("Swappee conflict in heat,room,index", heat_index, swapping_room, placed_inst[swapping_room].index(inst), 'inst', inst)
            for i, couples in enumerate(heat.getCouples()):
                if contestant in couples:
                    dup = True
                    break
            # if no conflicts with both lead and follow, make the swap with entry index i
            if math.fmod(index_iter, 2) == 1:
                if not dup:
                    index_2_swap = int((index_iter-1) / 2)
                    f = contestant
                    if [heat_index, swapping_room, index_2_swap] not in poachlist and (l not in poachcouples) and (f not in poachcouples): # Check this couple is not in the list already
                        poachlist.append([heat_index, swapping_room, index_2_swap])
                        poachcouples.append(l)
                        poachcouples.append(f)
                        break
                dup = False
            else:
                if not dup:
                    l = contestant
            index_iter += 1

    if (len(poachlist) + len(heat.getCouples()[roomid])) < acceptablecouples:
        print("Adding data back to dance df")
        for contestant in reversed(heat.getCouples()[roomid]):
            index = len(heat.getCouples()[roomid])-1
            tmp = heat.stealEntry(roomid, index)
            dance_df = pd.concat([dance_df, tmp])
        print("Backfilling", div)
        backfill(dance_df, div, heatlist, couples_per_floor, init.ev)
    else:
        print("Poaching from heats", div, poachlist)
        heatlen = len(heat.getRoster()[roomid])
        counter = 0
        while heatlen < acceptablecouples:
            for poach in poachlist:
                if heatlen == acceptablecouples:
                    break
                # poach = random.choice(poachlist)
                heat2poach = heatlist.getRostersList()[poach[0]]
                if len(heat2poach.getRoster()[poach[1]]) >= couples_per_floor and counter == 0:
                    print("Contestant", heat2poach.getCouples()[poach[1]][poach[2]], " stolen from heat, room", poach[0], poach[1])
                    tmp = heat2poach.stealEntry(poach[1], poach[2])
                    heat.addEntry(tmp, roomid)
                    poachlist.remove(poach)
                    heatlen += 1
                elif counter >= 1:
                    print("Contestant", heat2poach.getCouples()[poach[1]][poach[2]], " stolen from heat, room",poach[0], poach[1])
                    tmp = heat2poach.stealEntry(poach[1], poach[2])
                    heat.addEntry(tmp, roomid)
                    poachlist.remove(poach)
                    heatlen += 1
            counter += 1


def backfill(dance_df, div, heat_list, couples_per_floor, ev):
    # Todo: while df has entries
    #  loop over the heat list and find a heat that is < max
    #  if found check the meta data,
    #  check if contestant(s) is alreay in the heat
    #  loop over the inst of single and strike it if it is there, if there are no inst left move on

    list_index = 0
    heats = heat_list.getRostersList()
    heats_count = len(heats)
    while not dance_df.empty:
        for row, data in dance_df.iterrows():
            # Set columns to check
            if data["type id"] == "L":
                con_col = ["Lead Dancer #", "Lead Dancer #"]
            elif data["type id"] == "F":
                con_col = ["Follow Dancer #", "Follow Dancer #"]
            else:
                con_col = ["Lead Dancer #", "Follow Dancer #"]
            # Search the list
            placed = False
            list_index = 0
            while (list_index < heats_count):

                # Check if heat has this division
                if div not in heats[list_index].getDiv():
                    list_index += 1
                    continue
                # Check if Dancer is already in heat
                found = False
                for each in heats[list_index].getSingles():
                    if data[con_col[0]] in each or data[con_col[1]] in each:
                        found = True
                        break
                for each in heats[list_index].getInstructors():
                    if data[con_col[0]] in each or data[con_col[1]] in each:
                        print("While in backfill() Instructor # matches with " + data)
                        found = True
                        break
                for each in heats[list_index].getCouples():
                    if data[con_col[0]] in each or data[con_col[1]] in each:
                        found = True
                        break
                if found:
                    list_index += 1
                    break
                roomid = heats[list_index].getDiv().index(div)  # Note will only return
                # if couple place it
                if data["type id"] == "C":
                    heats[list_index].addEntry(data, roomid)
                    list_index += 1
                    break

                # If single need to check the instructors
                if data["type id"] == "L":
                    contestant_col = "Lead Dancer #"
                    inst_col = "Follow Dancer #"
                    inst_fname = "Follow First Name"
                    inst_lname = "Follow Last Name"
                elif data["type id"] == "F":
                    contestant_col = "Follow Dancer #"
                    inst_col = "Lead Dancer #"
                    inst_fname = "Lead First Name"
                    inst_lname = "Lead Last Name"
                else:
                    raise Exception("While running backfill() for", div, "Type id for", data, "is invalid")
                # Retrofit the instructor list data
                if type(data["Instructor Dancer #'s"]) == int:
                    data["Instructor Dancer #'s"] = [data["Instructor Dancer #'s"]]
                else:
                    pass
                    # tmp = [int(x) for x in data["Instructor Dancer #'s"].split(";")]

                for inst in data["Instructor Dancer #'s"]:
                    for each in heats[list_index].getInstructors():
                        if inst in each:
                            found = True
                            break
                    for each in heats[list_index].getCouples():
                        if inst in each:
                            found = True
                            break
                    for each in heats[list_index].getSingles():
                        if inst in each:
                            found = True
                            break
                    if (not found) and len(heats[list_index].getSingles()[roomid]) <= couples_per_floor:
                        # Set candidate to this single/inst match
                        candidate = data.to_frame().T
                        candidate = candidate.reset_index(drop=True)
                        instructor_data = init.df_inst[init.df_inst["Dancer #"] == inst].reset_index(drop=True)
                        candidate.loc[0, inst_col] = instructor_data.loc[0, "Dancer #"]
                        candidate.loc[0, inst_fname] = instructor_data.loc[0, "First Name"]
                        candidate.loc[0, inst_lname] = instructor_data.loc[0, "Last Name"]
                        heats[list_index].addEntry(candidate, roomid)
                        # Remove the placed candidate from the df, or -1 if multi-entry
                        dance_df.loc[dance_df.loc[:, contestant_col] == candidate.loc[0, contestant_col], ev] = candidate.loc[0, ev] - 1
                        break
                list_index += 1
                if not found:
                    break
        dance_df = dance_df[dance_df.loc[:, ev] != 0]

def checkheat(heat):
    for sing_room in heat.getSingles():
        for sing in sing_room:
            counter = 0
            for each in heat.getSingles():
                if sing in each:
                    counter += 1
                if each.count(sing) > 1:
                    raise Exception(" Multiple persons in this heat")


            for each in heat.getInstructors():
                if sing in each:
                    counter += 1
                if each.count(sing) > 1:
                    raise Exception(" Multiple persons in this heat")


            for each in heat.getCouples():
                if sing in each:
                    counter += 1
                if each.count(sing) > 1:
                    raise Exception(" Multiple persons in this heat")

            if counter > 1:
                raise Exception("Multiple Persons in this heat")

    for inst_room in heat.getInstructors():
        for inst in inst_room:
            counter = 0

            for each in heat.getSingles():
                if inst in each:
                    counter += 1
                if each.count(inst) > 1:
                    raise Exception(" Multiple persons in this heat")

            for each in heat.getInstructors():
                if inst in each:
                    counter += 1
                if each.count(inst) > 1:
                    raise Exception(" Multiple persons in this heat")

            for each in heat.getCouples():
                if inst in each:
                    counter += 1
                if each.count(inst) > 1:
                    raise Exception(" Multiple persons in this heat")
            if counter > 1:
                raise Exception("Multiple Persons in this heat")

    for coup_room in heat.getCouples():
        for coup in coup_room:
            counter = 0

            for each in heat.getSingles():
                if coup in each:
                    counter += 1
                if each.count(coup) > 1:
                    raise Exception(" Multiple persons in this heat")
            for each in heat.getInstructors():
                if coup in each:
                    counter += 1
                if each.count(coup) > 1:
                    raise Exception(" Multiple persons in this heat")
            for each in heat.getCouples():
                if coup in each:
                    counter += 1
                if each.count(coup) > 1:
                    raise Exception(" Multiple persons in this heat")

            if counter > 1:
                raise Exception("Multiple Persons in this heat")