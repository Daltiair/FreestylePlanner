from openpyxl.styles import PatternFill, Alignment
from openpyxl import load_workbook
import pandas as pd
from pandas import ExcelWriter
import os
import openpyxl
import init
from debug import checkheat
from output_methods import createParticipantSheets, appendParticipantSheet, clearParticipantSheetCounts
import shutil


def buildEvent(heats, eventName):
    # createParticipantSheets()
    rowindex = 1
    # file = os.getcwd() + eventName + '.xlsx'
    init.logString = ""
    print()
    print("Printing Heats")
    filepath = os.getcwd().replace('\src', "") + "\Output"
    shutil.rmtree(filepath)
    for each in heats:  # For each genre
        # wb.remove_sheet(wb.get_sheet_names()[0])
        for every in heats[each]:  # For every syllabus category
            events = list(heats[each][every].keys())  # Create list of keys
            if events == []:
                continue
            # Create a new directory
            filepath = os.getcwd().replace('\src', "") + "\Output" + "\\" + str(each) + "\\" + str(every) + "\\"
            # filepath = filepath.replace('\\', "/")
            try:
                os.makedirs(filepath)
            except:
                print("Filepath", filepath, "already exists")
            if heats[each][every] == {}:
                continue
            wb = openpyxl.Workbook()
            excelfile = eventName + '_' + str(each) + '_' + str(every) + '.xlsx'
            wb.save(filepath + excelfile)
            rowindex = 2
            for ev in events:  # Loop all events for this genre and syllabus
                init.ev = ev
                heatslist = heats[each][every][ev]  # Print all heats for the event
                if heatslist == []:
                    continue
                print(each, every, ev)
                for check in heatslist.getRostersList():
                    checkheat(check)
                wb.create_sheet(title=ev)
                wb.active = wb[ev]
                wb.active.page_setup.fitToWidth = 1
                wb.save(filepath+excelfile)
                rosters = heatslist.getRostersList()
                couples_per_floor = heatslist.getCouplesPerFloor()
                count = heatslist.getHeatCount()
                with ExcelWriter(filepath + excelfile, mode='a', if_sheet_exists="overlay") as writer:
                    for heat in rosters:  # Loop over all heats in the Heatlist obj
                        for roomid, room in enumerate(heat.getRoster()):  # For each ballroom print out a list, TODO: may need a -1
                            rowindex += 1
                            startingrow = rowindex
                            # print out each contestant, formatting df to relevant columns
                            for i, contestant in enumerate(room):
                                appendParticipantSheet(heat.getDiv()[roomid], every, ev, roomid, contestant, heat, heatslist)
                                if i == 0 and roomid == 0:
                                    contestant.to_excel(writer, sheet_name=ev, startrow=rowindex-1, columns=init.df_cols, index=False)
                                    rowindex += 1
                                else:
                                    contestant.to_excel(writer, sheet_name=ev, startrow=rowindex-1, columns=init.df_cols, index=False, header=False)
                                rowindex += 1
                            # add in blank rows if data < couples-per-floor
                            if len(room) < couples_per_floor:
                                for i in range(couples_per_floor - len(room)):
                                    rowindex += 1
                        rowindex += 1
                # go over the printed rows and highlight cells for easy identify
                rowindex -= 2
                wb = load_workbook(filename=filepath + excelfile)
                sheet = wb.get_sheet_by_name(ev)
                prev_floor = 1
                # print("Heats", len(heatslist.getRostersList()))
                for i in range(rowindex-1):
                    for col, aline in zip(init.excelcols, init.excelalignments):
                        sheet[col + str(i+1)].alignment = Alignment(horizontal=aline)
                rooms = heatslist.getFloors()
                roomid = 0
                roomindex = 0
                roommax = couples_per_floor + 3
                heatiter = 0
                for i in range(rowindex):
                    # Iterator data
                    index = i + 1
                    if roomindex < roommax:
                        roomindex += 1
                    else:
                        roomindex = 1
                        if roomid < (rooms - 1):
                            # print("New Room at", index)
                            roomid += 1
                        else:
                            # print("New Heat", heatiter +1 , "at", index)
                            roomid = 0
                            heatiter = heatiter + 1
                            # wb.save(filepath + excelfile)
                    # If a non contestant data row, add identifier rows, text and color
                    if roomindex == 1 or (roomid == 0 and roomindex == 2):
                        if roomid == 0:
                            roommax = couples_per_floor + 3
                            if roomindex == 1:
                                try:
                                    sheet[init.excelcols[0] + str(index)] = heatslist.getRostersList()[heatiter].getKey()
                                    sheet[init.excelcols[0] + str(index)].fill = PatternFill("solid", start_color="e6bee2")
                                except:
                                    print(index, heatiter, roomid, rowindex)
                            elif roomindex == 2:
                                # Replace lvl # with it's corresponding name EX: AB, FB AS-FS
                                # try:
                                div = heatslist.getRostersList()[heatiter].getDiv()[roomid]
                                # except:
                                #     print(index, heatiter, roomid, rowindex)
                                if div == []:
                                    sheet[init.excelcols[0] + str(index)] = 'Floor ' + str(roomid + 1) + " Not used"  # ILLEGAL_CHARACTERS_RE.sub(r'', str(room))
                                    sheet[init.excelcols[0] + str(index)].fill = PatternFill("solid", start_color="a9ebba")
                                    sheet.merge_cells(init.excelcols[0] + str(index) + ":" + init.excelcols[-1] + str(index))
                                    sheet[init.excelcols[0] + str(index)].alignment = Alignment(horizontal='left')
                                    continue
                                if div[0] == "S":
                                    for lev in init.lvl_conversion:
                                        if lev in div:
                                            div[div.index(lev)] = heatslist.getEventLvlSingles()[lev]
                                    for age in heatslist.getEventAgesSingles():
                                        if age in div:
                                            printagelist = heatslist.getEventAgesSingles().copy()
                                            printagelist.insert(0, 17)
                                            age_index = printagelist.index(div[div.index(age)])
                                            if age_index == printagelist[-1]:  # If the last bracket make it '86+'
                                                div[div.index(age)] = str(printagelist[age_index - 1] + 1) + "+"
                                            else:
                                                div[div.index(age)] = str(printagelist[age_index - 1] + 1) + "-" + str(printagelist[age_index])
                                if div[0] == "C":
                                    for lev in init.lvl_conversion:
                                        if lev in div:
                                            div[div.index(lev)] = heatslist.getEventLvlCouples()[lev]
                                    for age in heatslist.getEventAgesCouples():
                                        if age in div:
                                            printagelist = heatslist.getEventAgesCouples().copy()
                                            printagelist.insert(0, 17)
                                            age_index = printagelist.index(div[div.index(age)])
                                            if age_index == printagelist[-1]:  # If the last bracket make it '86+'
                                                div[div.index(age)] = str(printagelist[age_index - 1] + 1) + "+"
                                            else:
                                                div[div.index(age)] = str(printagelist[age_index - 1] + 1) + "-" + str(printagelist[age_index])
                                printstr = ""
                                for d in div:
                                    printstr = printstr + str(d) + ", "
                                sheet[init.excelcols[0] + str(index)] = 'Floor ' + str(roomid + 1) + " " + printstr[:-2]  # ILLEGAL_CHARACTERS_RE.sub(r'', str(room))
                                sheet[init.excelcols[0] + str(index)].fill = PatternFill("solid", start_color="a9ebba")
                        elif roomid > 0 and roomindex == 1:
                            roommax = couples_per_floor + 1
                            # Replace lvl # with it's corresponding name EX: AB, FB AS-FS
                            # try:
                            div = heatslist.getRostersList()[heatiter].getDiv()[roomid]
                            # except:
                            #     print(index, heatiter, roomid, rowindex)
                            if div == []:
                                sheet[init.excelcols[0] + str(index)] = 'Floor ' + str(roomid + 1) + " Not used"  # ILLEGAL_CHARACTERS_RE.sub(r'', str(room))
                                sheet[init.excelcols[0] + str(index)].fill = PatternFill("solid", start_color="a9ebba")
                                sheet.merge_cells(init.excelcols[0] + str(index) + ":" + init.excelcols[-1] + str(index))
                                sheet[init.excelcols[0] + str(index)].alignment = Alignment(horizontal='left')
                                continue
                            if div[0] == "S":
                                # for lev in init.lvl_conversion:
                                #     if lev in div:
                                #        div[div.index(lev)] = heatslist.getEventLvlSingles()[lev]
                                for age in heatslist.getEventAgesSingles():
                                    if age in div:
                                        printagelist = heatslist.getEventAgesSingles().copy()
                                        printagelist.insert(0, 17)
                                        age_index = printagelist.index(div[div.index(age)])
                                        if age_index == printagelist[-1]:  # If the last bracket make it '86+'
                                            div[div.index(age)] = str(printagelist[age_index - 1] + 1) + "+"
                                        else:
                                            div[div.index(age)] = str(printagelist[age_index - 1] + 1) + "-" + str(printagelist[age_index])
                            if div[0] == "C":
                                # for lev in init.lvl_conversion:
                                #     if lev in div:
                                #         div[div.index(lev)] = heatslist.getEventLvlCouples()[lev]
                                for age in heatslist.getEventAgesCouples():
                                    if age in div:
                                        printagelist = heatslist.getEventAgesCouples().copy()
                                        printagelist.insert(0, 17)
                                        age_index = printagelist.index(div[div.index(age)])
                                        if age_index == printagelist[-1]: # If the last bracket make it '86+'
                                            div[div.index(age)] = str(printagelist[age_index-1] + 1) + "+"
                                        else:
                                            div[div.index(age)] = str(printagelist[age_index-1] + 1) + "-" + str(printagelist[age_index])
                            printstr = ""
                            for d in div:
                                printstr = printstr + str(d) + ", "
                            sheet[init.excelcols[0] + str(index)] = 'Floor ' + str(roomid + 1) + " " + printstr[:-2] # ILLEGAL_CHARACTERS_RE.sub(r'', str(room))
                            sheet[init.excelcols[0] + str(index)].fill = PatternFill("solid", start_color="a9ebba")
                        sheet.merge_cells(init.excelcols[0]+str(index)+":"+init.excelcols[-1]+str(index))
                        sheet[init.excelcols[0] + str(index)].alignment = Alignment(horizontal='left')
                        # prev_floor = index
                        # wb.save(filepath+excelfile)
                        continue
                    # If Contestant Data header row
                    if roomid == 0 and roomindex == 3:
                        for col in init.excelcols:
                            sheet[col + str(index)].alignment = Alignment(horizontal='center')

                    # Color code the type of contestant
                    idcol = init.excelcols[0]
                    if sheet["A" + str(index)].value == None:
                        sheet[idcol+str(index)].fill = PatternFill("solid", start_color="0a0a0a")
                    if sheet[idcol+str(index)].value == "L":
                        sheet[idcol+str(index)].fill = PatternFill("solid", start_color="a2c2f5")
                    elif sheet[idcol+str(index)].value == "F":
                        sheet[idcol+str(index)].fill = PatternFill("solid", start_color="f2a7aa")
                    elif sheet[idcol+str(index)].value == "C":
                        sheet[idcol+str(index)].fill = PatternFill("solid", start_color="cc9ee8")

                    # level color code
                    levcol = init.excelcols[7]
                    if sheet[levcol + str(index)].value is not None:
                        # Bronze/NC
                        if sheet[levcol+str(index)].value[0] == "B" or sheet[levcol+str(index)].value[0] == "N":
                            if sheet[levcol+str(index)].value[1] == "1" or sheet[levcol+str(index)].value[1] == "2" or sheet[levcol+str(index)].value[1] == "C":
                                sheet[levcol+str(index)].fill = PatternFill("solid", start_color="eddcca")
                            if sheet[levcol+str(index)].value[1] == "3" or sheet[levcol+str(index)].value[1] == "4":
                                sheet[levcol+str(index)].fill = PatternFill("solid", start_color="f2ba7e")
                        # Silver
                        if sheet[levcol+str(index)].value[0] == "S":
                            if sheet[levcol+str(index)].value[1] == "1" or sheet[levcol+str(index)].value[1] == "2":
                                sheet[levcol+str(index)].fill = PatternFill("solid", start_color="e0dfde")
                            if sheet[levcol+str(index)].value[1] == "3" or sheet[levcol+str(index)].value[1] == "4":
                                sheet[levcol+str(index)].fill = PatternFill("solid", start_color="a39e99")
                        # Gold
                        if sheet[levcol+str(index)].value[0] == "G":
                            if sheet[levcol+str(index)].value[1] == "1" or sheet[levcol+str(index)].value[1] == "2":
                                sheet[levcol+str(index)].fill = PatternFill("solid", start_color="f0eec5")
                            if sheet[levcol+str(index)].value[1] == "3" or sheet[levcol+str(index)].value[1] == "4" or sheet[levcol+str(index)].value[1] == "B":
                                sheet[levcol+str(index)].fill = PatternFill("solid", start_color="f5ee71")
                # Set column dimensions
                for col, dim in zip(init.excelcols, init.exceldimensions):
                    sheet.column_dimensions[col].width = dim
            wb.save(filepath+excelfile)
            clearParticipantSheetCounts()
            print()
    print()
    print("Creating Personal Heat Sheets")
    # Loop over all participant sheets and print them
    participants = list(init.participantsheets.keys())
    for each in participants:
        df = init.participantsheets[each]["Data"]
        # Create a new directory
        # filepath = os.getcwd().replace('\src', "") + "/Output" + "/" + "Personal Sheets" + "/" + str(each) + "/"
        filepath = os.getcwd().replace('\src', "") + "/Output" + "/" + "Personal Sheets" + "/"
        # filepath = filepath.replace('\\', "/")
        try:
            os.makedirs(filepath)
        except:
            pass
            # print("Filepath", filepath, "already exists")
        wb = openpyxl.Workbook()
        excelfile = str(each) + '.xlsx'
        sheet = wb.active
        sheet.merge_cells(init.participantsheet_excelcols[0] + str(1) + ":" + init.participantsheet_excelcols[-1] + str(1))
        sheet[init.participantsheet_excelcols[0] + str(1)].value = str(each) + "  " + init.df_Dnum["First Name"].loc[init.df_Dnum["Dancer #"] == each].values[0] + ", " + init.df_Dnum["Last Name"].loc[init.df_Dnum["Dancer #"] == each].values[0]
        # Set column dimensions
        for col, dim in zip(init.participantsheet_excelcols, init.participantsheet_exceldimensions):
            sheet.column_dimensions[col].width = dim
        for i in range(df.shape[0]+1):
            index = i + 2
            for col, aline in zip(init.participantsheet_excelcols, init.participantsheet_excelalignments):
                sheet[col + str(index)].alignment = Alignment(horizontal=aline)
        wb.save(filepath + excelfile)
        with ExcelWriter(filepath + excelfile, mode='a', if_sheet_exists="overlay") as writer:
            df.to_excel(writer, sheet_name='Sheet', startrow=1, index=False)
    createZip()


def createZip():
    filename = "\\SortingErrors.txt"
    filepath = os.getcwd().replace('\src', "") + "\\Output"
    logfile = filepath + filename
    print(logfile)
    if init.logString != "":
        print("Creating Error Log file")
        with open(logfile, "w+") as f:
            f.write(init.logString)
    print("Creating zip file")
    zipfile = os.getcwd().replace('\src', "") + "\Output"+ "\Output_" + init.eventName
    outputfilepath = os.getcwd().replace('\src', "") + "/Output"
    # print(zipfile)
    shutil.make_archive(zipfile, 'zip', outputfilepath)


def makeHeatDict(genrelist, df_cat):
    heats = {}
    syllabus = ['Open', 'Closed']
    # For each genre, fill a dictionary key'd with genre, with a sub dict key'd by dance name with that genre
    # This will be used to hold all heats for that dance
    for each in genrelist:
        heats[each] = {}
        for every in syllabus:
            df_genre = df_cat[(df_cat.Genre == each) & (df_cat.Syllabus == every)]
            heats[each][every] = {}
            while not df_genre.empty:
                subdict = heats[each][every]
                subdict.update({df_genre["Dance"].iloc[0]: []})
                #heats.update(subdict.update({df_genre["Dance"].iloc[0]: []}))
                df_genre = df_genre.iloc[1:]
    return heats


def buildEventfast(heatlist, eventName):
    # createParticipantSheets()
    rowindex = 1
    # file = os.getcwd() + eventName + '.xlsx'
    print()
    print("Printing Heats")

    # Create a new directory
    filepath = os.getcwd().replace('\src', "") + "/Output" + "/Debug"+ "/"
    # filepath = filepath.replace('\\', "/")
    try:
        os.makedirs(filepath)
    except:
        print("Filepath", filepath, "already exists")
    # if heats[each][every] == {}:
    #     continue
    wb = openpyxl.Workbook()
    excelfile = eventName + '_' + "debug" + '.xlsx'
    wb.save(filepath + excelfile)
    rowindex = 2
    # heatslist = heats[each][every][ev]  # Print all heats for the event
    # if heatslist == []:
    #     continue
    # print(each, every, ev)
    wb.create_sheet(title=init.ev)
    wb.active = wb[init.ev]
    wb.active.page_setup.fitToWidth = 1
    wb.save(filepath+excelfile)
    rosters = heatlist.getRostersList()
    couples_per_floor = heatlist.getCouplesPerFloor()
    # count = heatslist.getHeatCount()
    with ExcelWriter(filepath + excelfile, mode='a', if_sheet_exists="overlay") as writer:
        for heat in rosters:  # Loop over all heats in the Heatlist obj
            for roomid, room in enumerate(heat.getRoster()):  # For each ballroom print out a list, TODO: may need a -1
                rowindex += 1
                startingrow = rowindex
                # if room == []:
                #     rowindex += 2
                # print out each contestant, formatting df to relevant columns
                for i, contestant in enumerate(room):
                    # appendParticipantSheet(heat.getDiv()[roomid], every, ev, roomid, contestant, heat, heatslist)
                    if i == 0 and roomid == 0:
                        contestant.to_excel(writer, sheet_name=init.ev, startrow=rowindex-1, columns=init.df_cols, index=False)
                        rowindex += 1
                    else:
                        contestant.to_excel(writer, sheet_name=init.ev, startrow=rowindex-1, columns=init.df_cols, index=False, header=False)
                    rowindex += 1
                # add in blank rows if data < couples-per-floor
                if len(room) < couples_per_floor:
                    for i in range(couples_per_floor - len(room)):
                        rowindex += 1
            rowindex += 1
    # go over the printed rows and highlight cells for easy identify
    rowindex -= 2
